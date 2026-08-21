# -*- coding: utf-8 -*-
"""
橫道圖 A3 橫向打印排版優化
--------------------------------------------------
問題診斷（原檔）：
  寬 135.7 字符 ≈ 251mm  vs A3橫向可打印寬 395mm  → 右側浪費 36%
  高 1393pt    ≈ 491mm  vs A3橫向可打印高 256mm  → 需 2 頁（超 92%）
  fitToWidth=0 → Excel 視為「不限頁寬」，自動縮放失效

解法：按 A3 橫向可打印區的寬高比（395:256）重配版面，
      令內容寬高比與紙張一致，fitToPage 時剛好鋪滿單頁。

換算基準：Calibri 11 標準字元寬 = 7px @96dpi = 1.852mm；1pt = 0.3528mm
"""
import shutil
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

SRC = "./backup_before_a3.xlsx"
OUT = "./示例裝修工程_施工進度橫道圖.xlsx"

MM_PER_CHAR = 1.852
MM_PER_PT = 0.3528

# A3 橫向可打印區（mm）；邊距取 left .59" right .4" top .91" bottom .71"
A3_W = 420.0 - (0.59 + 0.40) * 25.4      # ≈ 394.9
A3_H = 297.0 - (0.91 + 0.71) * 25.4      # ≈ 255.9

wb = openpyxl.load_workbook(SRC)
ws = wb.active

NUM_INFO = 6            # A..F 資訊列
GANTT_C0 = NUM_INFO + 1  # G
LAST_COL = ws.max_column  # X = 24
LAST_ROW = ws.max_row     # 39
GANTT_N = LAST_COL - NUM_INFO

# ── 1. 行高：壓縮至單頁可縮放範圍 ────────────────────────────
SEG_ROWS = [7, 14]                                   # 節段標題
TASK_ROWS = list(range(8, 14)) + list(range(15, 30))  # 21 條任務
row_h = {
    1: 32, 2: 22, 3: 6, 4: 22, 5: 24, 6: 24,   # 標題／資訊／月份／日期表頭
    30: 8, 31: 22, 32: 19, 33: 19,             # 圖例
    34: 8, 35: 22, 36: 19, 37: 19, 38: 19, 39: 19,  # 關鍵路徑說明
}
for r in SEG_ROWS:
    row_h[r] = 26
for r in TASK_ROWS:
    row_h[r] = 24

for r, h in row_h.items():
    ws.row_dimensions[r].height = h

total_pt = sum(row_h.get(r, 15) for r in range(1, LAST_ROW + 1))

# ── 2. 列寬：按紙張寬高比反推總寬，鋪滿 A3 橫向 ──────────────
# 需求：W*MM_PER_CHAR*s = A3_W 且 total_pt*MM_PER_PT*s = A3_H
scale = A3_H / (total_pt * MM_PER_PT)
target_chars = A3_W / MM_PER_CHAR / scale

info_w = {1: 5.0, 2: 62.0, 3: 7.0, 4: 13.0, 5: 13.0, 6: 13.0}
info_sum = sum(info_w.values())
gantt_w = round((target_chars - info_sum) / GANTT_N, 2)

for c, w in info_w.items():
    ws.column_dimensions[get_column_letter(c)].width = w
for i in range(GANTT_N):
    ws.column_dimensions[get_column_letter(GANTT_C0 + i)].width = gantt_w

total_chars = info_sum + gantt_w * GANTT_N

# ── 3. 打印設定：A3 橫向、單頁、置中 ─────────────────────────
if ws.sheet_properties.pageSetUpPr is None:
    ws.sheet_properties.pageSetUpPr = PageSetupProperties()
ws.sheet_properties.pageSetUpPr.fitToPage = True

ps = ws.page_setup
ps.paperSize = 8            # 8 = A3
ps.orientation = "landscape"
ps.fitToWidth = 1           # 原為 0 → Excel 當「不限」，縮放失效
ps.fitToHeight = 1
ps.scale = None             # fitToPage 生效時必須留空
ps.horizontalDpi = 600
ps.verticalDpi = 600

ws.print_area = f"A1:{get_column_letter(LAST_COL)}{LAST_ROW}"
ws.print_title_rows = "1:6"
ws.print_options.horizontalCentered = True
ws.print_options.verticalCentered = False
ws.print_options.gridLines = False   # 交付文件不印網格線
ws.print_options.headings = False    # 不印行號列標

ws.page_margins.left = 0.59
ws.page_margins.right = 0.40
ws.page_margins.top = 0.91
ws.page_margins.bottom = 0.71
ws.page_margins.header = 0.30
ws.page_margins.footer = 0.30

# 頁眉頁腳（原為 &F 檔名 + "Confidential"，不適合投標文件）
ws.oddHeader.left.text = "示例裝修工程"
ws.oddHeader.center.text = ""
ws.oddHeader.right.text = "施工進度橫道圖"
ws.oddFooter.left.text = "示例承建商"
ws.oddFooter.center.text = ""
ws.oddFooter.right.text = "第 &P 頁 / 共 &N 頁"
for part in (ws.oddHeader.left, ws.oddHeader.right,
             ws.oddFooter.left, ws.oddFooter.right):
    part.size = 8
    part.font = "微軟正黑體"

ws.sheet_view.showGridLines = False
ws.sheet_view.zoomScale = 70

wb.save(OUT)

pw = total_chars * MM_PER_CHAR
ph = total_pt * MM_PER_PT
print("=== A3 橫向排版結果 ===")
print(f"內容尺寸      : {pw:.1f} x {ph:.1f} mm  (寬高比 {pw/ph:.3f})")
print(f"A3橫向可打印區: {A3_W:.1f} x {A3_H:.1f} mm  (寬高比 {A3_W/A3_H:.3f})")
print(f"預期縮放      : 寬 {A3_W/pw*100:.1f}%  高 {A3_H/ph*100:.1f}%  → Excel 取 {min(A3_W/pw, A3_H/ph)*100:.1f}%")
print(f"列寬          : 資訊列合計 {info_sum} 字符 / 甘特 {GANTT_N} 列 × {gantt_w} 字符")
print(f"行高總計      : {total_pt} pt  (任務行 24pt × {len(TASK_ROWS)})")
print(f"甘特列實印寬  : {gantt_w*MM_PER_CHAR*min(A3_W/pw, A3_H/ph):.1f} mm/列（原 {3.0*MM_PER_CHAR:.1f} mm）")
print(f"輸出          : {OUT}")
