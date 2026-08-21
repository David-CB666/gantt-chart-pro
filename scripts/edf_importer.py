#!/usr/bin/env python3
"""
EDF Importer v3.0 — 智能、自適應 Excel → 甘特圖 JSON 轉換器
============================================================
設計目標：面對唔同格式嘅進度表（模版 A / B / C 甚至未來新格式），
          都能自動識別出「表頭 / 列頭 / 內容 / 關鍵詞」，毋須逐個硬編。

核心特性
--------
1. 表頭行自動偵測：掃描前 N 行，按關鍵詞命中評分選出表頭行（唔假設一定喺第 1 行）。
2. 列頭 → 語義欄位 映射：seq/code/name/start/end/duration/predecessor/category/material/remark，
   全部由 recognition_config.json 嘅關鍵詞字典驅動，新增格式只加關鍵詞。
3. 多格式日期解析：文本(含中文+時分)、datetime 對象、Excel 序列號，統一處理。
4. 工序期雜質容錯：'160 d?'、'37 d'、'5 d' 都用正則抽取首個整數。
5. 雙模式前置依賴解析：
     - 行號模式（模版 A / C）：前置 = Excel 行號（可帶 SS/FF/SF/FS 關係尾碼）
     - 代碼模式（模版 B）：前置 = 任務代碼（如 '1.2'、'2.1; 2.2'）
   自動按前置值形態判定模式。
6. 通用元數據抽取：掃描全表，將「工程名稱 / 工程編號 / 承建商 / 總工期 / 編制時間 ...」
   等鍵值對抽取到 project 欄位，唔受側欄位置限制。
7. 關鍵詞透明化：--report 會產出識別報告，列出表頭、列映射、元數據、每個任務嘅
   自動分類及命中關鍵詞，方便核對。

用法
----
  python edf_importer.py <EDF.xlsx> [-o output.json] [--project 名] [--subtitle 副標]
                                    [--sheet 表名] [--report] [--profile 名]

由 gen_gantt.py 調用（保持向下兼容，return dict）：
  from edf_importer import import_edf
  task_config = import_edf("EDF.xlsx", project_name="我的項目")
  # 如需識別報告：
  config, report = import_edf("EDF.xlsx", return_report=True)
"""

import json
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, date, timedelta
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass


# ================================================================
#  內置預設（當 recognition_config.json 唔喺時用，保證腳本可獨立執行）
# ================================================================

DEFAULT_FIELD_PATTERNS: Dict[str, List[str]] = {
    "seq": ["序號", "項次", "編號", "項號", "No", "no", "NO", "Num", "#", "seq", "序号", "项号"],
    "code": ["任務代碼", "任务代码", "工序編號", "工序编号", "代碼", "code", "Code", "WBS"],
    "name": ["施工內容", "施工項目", "項目名稱", "項目", "任務名稱", "任務描述", "工序",
             "工程內容", "內容", "工作內容", "名稱", "name", "Name", "施工内容", "项目名称"],
    "start": ["開始日期", "開始時間", "開始", "開工日期", "開工", "開始", "Start", "start", "SD"],
    "end": ["完成日期", "完成時間", "完成", "完工日期", "完工", "竣工日期", "結束", "End", "end", "ED"],
    "duration": ["工期", "持續天數", "持續", "天數", "工作天", "日數", "天", "Duration", "dur"],
    "predecessor": ["前置任務", "前置", "先行任務", "先決任務", "依賴", "前置任务", "Predecessor", "pred"],
    "material": ["材料訂貨期", "材料订货期", "材料到場", "訂貨期", "材料", "Material"],
    "remark": ["備註", "備注", "說明", "Remark", "Note", "備考", "备注"],
    "category": ["類別", "專業", "category", "type", "階段", "阶段", "Phase"],
}

DEFAULT_META_PATTERNS: Dict[str, List[str]] = {
    "project_name": ["工程名稱", "工程名称", "項目名稱", "项目名称", "project name", "project_name", "project"],
    "project_id": ["工程編號", "工程编号", "項目編號", "项目编号", "詢價編號", "询价编号", "project id", "rfq"],
    "contractor": ["承建商", "承辦商", "施工單位", "施工单位", "承包商", "contractor"],
    "prepare_date": ["編制時間", "编制时间", "編制日期", "编制日期", "prepare date", "prepare_date"],
    "total_duration": ["總工期", "总工期", "total duration"],
    "bid_date": ["投標日期", "投标日期", "bid date"],
    "location": ["項目地點", "项目地点", "地點", "地点", "location"],
}

DEFAULT_CATEGORY_KEYWORDS: Dict[str, List[str]] = {
    "A": ["準備", "開工", "交底", "勘察", "測量", "放線", "採購", "訂貨", "進場", "籌備", "設計", "前期"],
    "B": ["拆卸", "拆除", "清拆", "打拆", "批盪", "批灰", "抹灰", "防水", "結構", "加固", "鋼筋", "土建", "砌牆"],
    "C": ["天花", "牆身", "牆面", "地台", "地面", "地板", "地磚", "瓷磚", "油漆", "門", "窗", "木工", "飾面", "裝修"],
    "D": ["電氣", "配電", "電箱", "空調", "通風", "消防", "給水", "排水", "弱電", "監控", "電梯", "照明"],
    "E": ["安裝", "家具", "潔具", "五金", "扶手", "欄杆", "窗簾", "標識", "家電", "軟裝"],
    "F": ["測試", "調試", "驗收", "竣工", "交付", "整改", "保修", "培訓", "聯調"],
    "G": ["其他", "雜項", "預留", "應急", "政府", "報批", "審圖"],
}

EXCEL_EPOCH = datetime(1899, 12, 30)
CN_DATE_RE = re.compile(r"(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日")
DEFAULT_DATE_FORMATS = [
    "%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%m-%d-%Y",
    "%Y.%m.%d", "%Y年%m月%d日", "%d.%m.%Y",
]


# ================================================================
#  配置加載
# ================================================================

def _script_dir() -> str:
    return os.path.dirname(os.path.abspath(__file__))


def load_recognition_config(path: Optional[str] = None) -> dict:
    """載入 recognition_config.json；唔到就用內置預設。"""
    if path is None:
        path = os.path.join(_script_dir(), "recognition_config.json")
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                cfg = json.load(f)
            print(f"  [配置] 已載入識別字典: {os.path.basename(path)}")
            return cfg
        except Exception as e:
            print(f"  [!] 讀取識別配置失敗 ({e})，使用內置預設。")
    else:
        print("  [配置] 未找到 recognition_config.json，使用內置預設。")
    return {
        "fields": DEFAULT_FIELD_PATTERNS,
        "metadata": DEFAULT_META_PATTERNS,
        "category_keywords": DEFAULT_CATEGORY_KEYWORDS,
        "matching": {"exact_first": True, "fuzzy_contains": True, "min_fuzzy_len": 2, "meta_key_max_len": 12},
        "date": {"excel_serial_min": 20000, "excel_serial_max": 60000,
                 "valid_year_min": 2000, "valid_year_max": 2040, "formats": DEFAULT_DATE_FORMATS},
        "duration": {"suffixes": ["d", "天", "day", "days", "日"]},
        "predecessor": {"separators": [",", "，", ";", "；", "、", "/", " "],
                        "relation_suffixes": ["SS", "FF", "SF", "FS"],
                        "row_number_regex": r"^\d+(?:SS|FF|SF|FS)?$"},
        "header_scan_rows": 6,
    }


# ================================================================
#  通用工具
# ================================================================

# 繁簡中性化：將常見繁體字元摺疊為簡體，令表頭無論用繁/簡都能匹配。
# 新增格式時若用到其他繁簡異體字，在此補一對即可。
_FOLD_MAP = {
    "務": "务", "項": "项", "號": "号", "內": "内", "稱": "称", "開": "开", "動": "动",
    "計": "计", "製": "制", "編": "编", "處": "处", "單": "单", "總": "总", "點": "点",
    "細": "细", "規": "规", "圖": "图", "準": "准", "裝": "装", "備": "备", "註": "注",
    "說": "说", "類": "类", "專": "专", "統": "统", "測": "测", "線": "线", "門": "门",
    "閉": "闭", "關": "关", "電": "电", "氣": "气", "機": "机", "來": "来", "進": "进",
    "場": "场", "書": "书", "飛": "飞", "業": "业", "當": "当", "時": "时", "間": "间",
    "題": "题", "館": "馆", "庫": "库", "級": "级", "實": "实", "對": "对", "應": "应",
    "體": "体", "術": "术", "認": "认", "證": "证", "變": "变", "檢": "检", "驗": "验",
    "導": "导", "義": "义", "網": "网", "頁": "页", "設": "设", "許": "许", "話": "话",
    "語": "语", "辦": "办", "報": "报", "據": "据", "擇": "择", "澤": "泽", "絕": "绝",
    "結": "结", "約": "约", "紅": "红", "紙": "纸", "維": "维", "護": "护", "評": "评",
    "親": "亲", "輸": "输", "輔": "辅", "轉": "转", "軟": "软", "車": "车", "載": "载",
    "過": "过", "運": "运", "還": "还", "這": "这", "適": "适", "質": "质", "責": "责",
    "費": "费", "資": "资", "購": "购", "貨": "货", "軍": "军", "農": "农", "師": "师",
    "節": "节", "與": "与", "構": "构", "鋼": "钢", "鐵": "铁", "錯": "错", "銀": "银",
    "銅": "铜", "鍵": "键", "鎖": "锁", "鏡": "镜", "長": "长", "東": "东", "馬": "马",
    "鳥": "鸟", "魚": "鱼", "鹵": "卤", "麥": "麦", "黃": "黄", "齊": "齐", "濕": "湿",
    "溫": "温", "禮": "礼", "視": "视", "覺": "觉", "觸": "触", "議": "议", "詞": "词",
    "試": "试", "誤": "误", "調": "调", "請": "请", "諸": "诸", "讀": "读", "負": "负",
    "載": "载", "週": "周", "歷": "历", "歲": "岁", "縣": "县", "醫": "医", "顯": "显",
}
_FOLD_TABLE = str.maketrans(_FOLD_MAP)


def _normalize(s: Any) -> str:
    if s is None:
        return ""
    t = str(s).strip().lower()
    if _FOLD_TABLE:
        t = t.translate(_FOLD_TABLE)
    return t


def _valid_year(d: Optional[date], cfg: dict) -> bool:
    if d is None:
        return False
    lo = cfg.get("date", {}).get("valid_year_min", 2000)
    hi = cfg.get("date", {}).get("valid_year_max", 2040)
    return lo <= d.year <= hi


# ================================================================
#  欄位 / 元數據 關鍵詞匹配
# ================================================================

def _match_one(header: str, patterns: List[str], cfg: dict) -> bool:
    """判斷一個表頭文字是否命中某欄位嘅關鍵詞集合。"""
    norm = _normalize(header)
    if not norm:
        return False
    m = cfg.get("matching", {})
    # 精確
    for pat in patterns:
        if _normalize(pat) == norm:
            return True
    # 模糊包含
    if m.get("fuzzy_contains", True):
        min_len = m.get("min_fuzzy_len", 2)
        for pat in patterns:
            pn = _normalize(pat)
            if len(pn) < min_len:
                continue
            if pn in norm or norm in pn:
                return True
    return False


def detect_columns(headers: List[str], field_patterns: Dict[str, List[str]], cfg: dict) -> Dict[str, int]:
    """
    自動檢測列頭 → field → column_index(0-based) 映射。
    優先序：按 field_patterns 字典順序，先命中者得，避免一列多義。
    """
    mapping: Dict[str, int] = {}
    for idx, h in enumerate(headers):
        if h is None:
            continue
        for field, patterns in field_patterns.items():
            if field in mapping:
                continue
            if _match_one(str(h), patterns, cfg):
                mapping[field] = idx
                break
    return mapping


def detect_metadata(ws, meta_patterns: Dict[str, List[str]], cfg: dict) -> Dict[str, str]:
    """
    通用元數據抽取：掃描全表每個單元格，
    若其文字命中元數據關鍵詞（且似標籤，非長任務名），
    則由「右→下→左」鄰居取數值。
    """
    meta: Dict[str, str] = {}
    maxr = ws.max_row
    maxc = ws.max_column
    mcfg = cfg.get("matching", {})
    max_key_len = mcfg.get("meta_key_max_len", 12)

    def neighbor(r: int, c: int) -> Optional[str]:
        for nr, nc in ((r, c + 1), (r + 1, c), (r, c - 1)):
            if 1 <= nr <= maxr and 1 <= nc <= maxc:
                v = ws.cell(row=nr, column=nc).value
                if v is not None and str(v).strip():
                    return str(v).strip()
        return None

    for r in range(1, maxr + 1):
        for c in range(1, maxc + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            norm = _normalize(v)
            if not norm:
                continue
            # 標籤應該短；長文字大概率是內容，跳過模糊匹配（精確仍可）
            is_short = len(str(v).strip()) <= max_key_len
            matched_field = None
            for fld, pats in meta_patterns.items():
                for pat in pats:
                    pn = _normalize(pat)
                    if pn == norm:
                        matched_field = fld
                        break
                    # 模糊：只允許「關鍵詞包含於單元格」(pattern ⊂ cell)，
                    # 禁反向 (cell ⊂ pattern)，否則 '工期' 會誤中 '總工期' 而偷走鄰居值。
                    if is_short and len(pn) >= 2 and pn in norm:
                        matched_field = fld
                        break
                if matched_field:
                    break
            if not matched_field:
                continue
            val = neighbor(r, c)
            if val:
                # 唔覆蓋已有值（首次命中優先）
                meta.setdefault(matched_field, val)
    return meta


# ================================================================
#  日期 / 工序期 / 前置 解析
# ================================================================

def parse_date_value(val: Any, cfg: dict) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val

    # Excel 序列號
    if isinstance(val, (int, float)):
        serial = float(val)
        dmin = cfg.get("date", {}).get("excel_serial_min", 20000)
        dmax = cfg.get("date", {}).get("excel_serial_max", 60000)
        if dmin <= serial <= dmax:
            # 先用 openpyxl 嘅 epoch 換算
            try:
                d = from_excel(serial)
                if _valid_year(d, cfg):
                    return d.date()
            except Exception:
                pass
            # 回退：EXCEL_EPOCH
            try:
                d = EXCEL_EPOCH + timedelta(days=serial)
                if _valid_year(d, cfg):
                    return d.date()
            except Exception:
                pass
        return None

    s = str(val).strip()
    if not s:
        return None

    # 中文日期（含時分亦無妨，正則只抽年月日）
    m = CN_DATE_RE.search(s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except Exception:
            pass

    for fmt in cfg.get("date", {}).get("formats", DEFAULT_DATE_FORMATS):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s).date()
    except Exception:
        pass
    return None


def parse_duration(val: Any, cfg: dict) -> Optional[int]:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        try:
            return int(float(val))
        except Exception:
            return None
    s = str(val).strip().lower()
    # 抽取首個整數，自動容錯 '160 d?'、'37 d'、'5 d'、'60天'
    m = re.search(r"(\d+)", s)
    if m:
        try:
            return int(m.group(1))
        except Exception:
            return None
    return None


def parse_predecessors(val: Any, cfg: dict) -> List[Tuple[str, str]]:
    """
    解析前置任務，返回 [('row'|'code', value), ...]。
    自動判定模式：純數字(可帶關係尾碼) → 行號；含 '.'/字母 → 代碼。
    """
    if val is None:
        return []
    if isinstance(val, (int, float)):
        return [("row", str(int(val)))]
    s = str(val).strip()
    if not s:
        return []
    seps = cfg.get("predecessor", {}).get("separators", [",", "，", ";", "；", "/", " "])
    for sep in seps:
        s = s.replace(sep, ",")
    parts = [p.strip() for p in s.split(",") if p.strip()]
    result: List[Tuple[str, str]] = []
    rel_suffixes = cfg.get("predecessor", {}).get("relation_suffixes", ["SS", "FF", "SF", "FS"])
    for p in parts:
        rel = None
        for suf in rel_suffixes:
            if p.endswith(suf):
                rel = suf
                p = p[: -len(suf)].strip()
                break
        # 模式判定：純數字 → 行號；否則 → 代碼
        if re.fullmatch(r"\d+", p):
            result.append(("row", p))
        else:
            result.append(("code", p))
    return result


# ================================================================
#  自動分類（關鍵詞 → A~G）
# ================================================================

def auto_classify(task_name: str, cat_keywords: Dict[str, List[str]]) -> Tuple[str, str]:
    """返回 (類別, 命中關鍵詞)。匹配唔到返 ('A', '')。"""
    if not task_name:
        return "A", ""
    for cat in ["A", "B", "C", "D", "E", "F", "G"]:
        for kw in cat_keywords.get(cat, []):
            if kw in task_name:
                return cat, kw
    return "A", ""


# ================================================================
#  表頭行自動偵測
# ================================================================

def _nearest(keys: List[int], target: int) -> Optional[int]:
    """在已排序嘅已解析行號中，搵離 target 最近嘅一個（用於依賴指向被跳過嘅行）。"""
    if not keys:
        return None
    keys = sorted(keys)
    pos = min(range(len(keys)), key=lambda i: abs(keys[i] - target))
    return keys[pos]


def detect_header_row(ws, field_patterns: Dict[str, List[str]],
                      meta_patterns: Dict[str, List[str]], cfg: dict) -> int:
    """掃描前 N 行，評分選出表頭行。返 1-based 行號。"""
    scan = int(cfg.get("header_scan_rows", 6))
    scan = max(1, min(scan, ws.max_row))
    best_row, best_score = 1, -1
    for r in range(1, scan + 1):
        score = 0
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            norm = _normalize(v)
            if not norm:
                continue
            # 欄位關鍵詞（權重 2）
            for pats in field_patterns.values():
                if _match_one(str(v), pats, cfg):
                    score += 2
                    break
            else:
                # 元數據關鍵詞（權重 1）
                for pats in meta_patterns.values():
                    if _match_one(str(v), pats, cfg):
                        score += 1
                        break
        if score > best_score:
            best_score, best_row = score, r
    return best_row


# ================================================================
#  主導入引擎
# ================================================================

def import_edf(
    edf_path: str,
    project_name: str = "",
    subtitle: str = "",
    output_json: Optional[str] = None,
    sheet_name: Optional[str] = None,
    config_path: Optional[str] = None,
    profile: Optional[str] = None,
    return_report: bool = False,
) -> Any:
    """
    從 EDF.xlsx 導入任務數據，返標準 gantt_config dict。
    若 return_report=True，返 (config, report_dict)。
    """
    if not os.path.exists(edf_path):
        raise FileNotFoundError(f"找不到文件: {edf_path}")

    rcfg = load_recognition_config(config_path)
    field_patterns = rcfg.get("fields", DEFAULT_FIELD_PATTERNS)
    meta_patterns = rcfg.get("metadata", DEFAULT_META_PATTERNS)
    cat_keywords = rcfg.get("category_keywords", DEFAULT_CATEGORY_KEYWORDS)

    wb = load_workbook(edf_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    # ---- 1. 表頭行偵測 ----
    header_row = detect_header_row(ws, field_patterns, meta_patterns, rcfg)
    print(f"  表頭行偵測: 第 {header_row} 行")

    # ---- 2. 讀取表頭，映射欄位 ----
    headers: List[str] = []
    for col in range(1, ws.max_column + 1):
        h = ws.cell(row=header_row, column=col).value
        headers.append(str(h).strip() if h is not None else "")
    print(f"  檢測到 {len(headers)} 列: {headers}")

    mapping = detect_columns(headers, field_patterns, rcfg)
    print(f"  列映射: {mapping}")

    if "name" not in mapping:
        raise ValueError(f"未檢測到「項目名稱/施工內容」列（表頭行={header_row}）。請檢查 EDF 表頭。")
    if "start" not in mapping:
        raise ValueError("未檢測到「開始日期」列。")
    has_end = "end" in mapping
    has_duration = "duration" in mapping
    if not has_end and not has_duration:
        raise ValueError("未檢測到「完成日期」或「工期」列，至少需要一個。")

    has_code = "code" in mapping

    # ---- 3. 通用元數據抽取 ----
    meta = detect_metadata(ws, meta_patterns, rcfg)
    # project_name 清洗（模板 B 含 '| 60日曆天 | ...'）
    raw_project_name = meta.get("project_name", "")
    if raw_project_name:
        raw_project_name = raw_project_name.split("|")[0].strip()
    print(f"  元數據: {meta}")

    # ---- 4. 逐行讀取任務 ----
    raw_tasks: List[dict] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        task: dict = {}
        for field, col_idx in mapping.items():
            task[field] = ws.cell(row=row_idx, column=col_idx + 1).value
        task["_excel_row"] = row_idx
        raw_tasks.append(task)

    # ---- 5. 解析日期 / 工序期 / 前置 / 分類 ----
    parsed_tasks: List[dict] = []
    skipped: List[dict] = []
    code_values: List[Optional[str]] = []  # 與 parsed_tasks 同位，存任務代碼

    for t in raw_tasks:
        raw_name = t.get("name")
        if raw_name is None or str(raw_name).strip() == "":
            continue  # 空行（含 None / 空白）
        name = str(raw_name).strip()
        start_date = parse_date_value(t.get("start"), rcfg)
        if start_date is None:
            skipped.append({"row": t["_excel_row"], "name": name, "reason": "無效開始日期"})
            continue

        end_date = parse_date_value(t.get("end"), rcfg)
        duration = parse_duration(t.get("duration"), rcfg)

        if duration is None and end_date:
            duration = (end_date - start_date).days + 1
        if duration is None:
            duration = 1
        if end_date is None and duration:
            end_date = start_date + timedelta(days=duration - 1)

        seq = t.get("seq")
        if seq is not None:
            try:
                seq = int(float(str(seq)))
            except (ValueError, TypeError):
                seq = str(seq).strip()

        code_val = str(t.get("code", "")).strip() if t.get("code") else ""
        deps_raw = parse_predecessors(t.get("predecessor"), rcfg)

        # 分類：欄位 > 自動關鍵詞
        raw_cat = t.get("category")
        if raw_cat and str(raw_cat).strip():
            cat = str(raw_cat).strip()[:1].upper()
            matched_kw = "(欄位指定)"
        else:
            cat, matched_kw = auto_classify(name, cat_keywords)

        parsed_tasks.append({
            "seq": seq,
            "name": name,
            "start_date": start_date,
            "end_date": end_date,
            "duration": duration,
            "deps_raw": deps_raw,
            "category": cat,
            "matched_kw": matched_kw,
            "_excel_row": t["_excel_row"],
            "material": str(t.get("material", "")).strip() if t.get("material") else "",
            "remark": str(t.get("remark", "")).strip() if t.get("remark") else "",
        })
        code_values.append(code_val)

    if not parsed_tasks:
        raise ValueError("未解析到任何有效任務。")

    cat_counts = Counter(t["category"] for t in parsed_tasks)
    cat_summary = ", ".join(f"{c}:{n}" for c, n in sorted(cat_counts.items()))
    print(f"  成功解析 {len(parsed_tasks)} 個任務，自動分類: {cat_summary}")
    if skipped:
        print(f"  跳過 {len(skipped)} 行（無有效開始日期，多為備註/等候期）")

    # ---- 6. 項目日期範圍 ----
    project_start = min(t["start_date"] for t in parsed_tasks)
    project_end = max(t["end_date"] for t in parsed_tasks)
    num_days = (project_end - project_start).days + 1

    for t in parsed_tasks:
        t["start_day"] = (t["start_date"] - project_start).days

    # ---- 7. 構建 sections（按類別分組） ----
    categories: Dict[str, List[dict]] = {}
    cat_order: List[str] = []
    for t in parsed_tasks:
        cat = t["category"]
        if cat not in categories:
            categories[cat] = []
            cat_order.append(cat)
        categories[cat].append(t)

    stage_colors = {
        "A": "2E86AB", "B": "81C784", "C": "FFD54F",
        "D": "E57373", "E": "9575CD", "F": "64B5F6", "G": "BFCF91",
    }
    cat_labels = {
        "A": "開工準備", "B": "拆卸及基層處理", "C": "裝修主體施工",
        "D": "電氣及設備安裝", "E": "安裝及收尾工程", "F": "測試驗收", "G": "附屬工程",
    }

    sections = []
    task_id = 0
    id_mapping: Dict[Tuple, int] = {}
    code_map: Dict[str, int] = {}
    for cat in cat_order:
        group = categories[cat]
        section_tasks = []
        for t in group:
            task_id += 1
            id_mapping[(t["seq"], t["name"])] = task_id
            # 代碼模式：建 code → id 映射
            cv = t.get("_code") if "_code" in t else None
            section_tasks.append({
                "id": task_id,
                "name": t["name"],
                "duration": t["duration"],
                "start_day": t["start_day"],
                "deps": [],
                "category": cat,
                "material": t["material"],
                "remark": t["remark"],
                "milestone": False,
                "progress": 0,
            })
        sections.append({"title": cat_labels.get(cat, f"階段 {cat}"), "tasks": section_tasks})

    # 建立 code_map（用 code_values 對應 parsed_tasks）
    for t, cv in zip(parsed_tasks, code_values):
        if cv:
            code_map[cv] = id_mapping.get((t["seq"], t["name"]))

    # ---- 8. 依賴映射（雙模式） ----
    row_to_task: Dict[int, dict] = {t["_excel_row"]: t for t in parsed_tasks}
    unresolved: List[dict] = []
    fallback_deps: List[dict] = []
    for sec in sections:
        for task in sec["tasks"]:
            # 搵返原始 deps_raw
            orig = None
            for t in parsed_tasks:
                if t["name"] == task["name"] and id_mapping.get((t["seq"], t["name"])) == task["id"]:
                    orig = t
                    break
            if not orig or not orig["deps_raw"]:
                continue
            mapped = []
            for mode, val in orig["deps_raw"]:
                if mode == "row":
                    dep_row = int(val)
                    o = row_to_task.get(dep_row)
                    if o is None:
                        # 指向嘅行被跳過（等候期/摘要行等），回退搵最近已解析行
                        nr = _nearest(list(row_to_task.keys()), dep_row)
                        if nr is not None:
                            o = row_to_task[nr]
                            fallback_deps.append(
                                {"task": task["name"], "ref": f"row#{val}", "resolved_to": f"row#{nr}"})
                    if o is not None:
                        mid = id_mapping.get((o["seq"], o["name"]))
                        if mid and mid != task["id"]:
                            mapped.append(mid)
                        elif mid == task["id"]:
                            pass  # 自指，忽略
                        else:
                            unresolved.append({"task": task["name"], "ref": f"row#{val}"})
                    else:
                        unresolved.append({"task": task["name"], "ref": f"row#{val}"})
                else:  # code
                    mid = code_map.get(val)
                    if mid and mid != task["id"]:
                        mapped.append(mid)
                    else:
                        unresolved.append({"task": task["name"], "ref": f"code#{val}"})
            task["deps"] = mapped

    # ---- 9. 項目名稱解析（優先序：--project > 元數據 > 首個任務 > 檔名） ----
    first_task_name = parsed_tasks[0]["name"] if parsed_tasks else ""
    if not project_name:
        if raw_project_name:
            project_name = raw_project_name
        elif first_task_name:
            project_name = first_task_name
        else:
            project_name = os.path.splitext(os.path.basename(edf_path))[0]

    # 工程名稱透明化：紀錄元數據名 與 首個任務標題，供用戶核對
    # （唔自動斷言邊個啱，因為工序/分節標題常含『工程』等字眼）。
    name_note = ""
    if raw_project_name and first_task_name and raw_project_name != first_task_name:
        name_note = (f"工程名稱取自元數據『工程名稱』={raw_project_name}；"
                     f"首個任務標題={first_task_name}。如有落差可用 --project 覆蓋。")

    # ---- 10. 構建最終配置 ----
    config = {
        "project": {
            "name": project_name,
            "subtitle": subtitle or f"{project_name} 施工進度橫道圖",
            "project_id": meta.get("project_id", ""),
            "contractor": meta.get("contractor", ""),
            "location": meta.get("location", ""),
            "prepare_date": meta.get("prepare_date", ""),
            "total_duration": meta.get("total_duration", ""),
            "start_date": project_start.isoformat(),
            "num_days": num_days,
        },
        "sections": sections,
        "critical_path": [],
        "legend": [
            {"color": stage_colors.get(c, "AAAAAA"), "label": cat_labels.get(c, c)}
            for c in cat_order
        ],
        "notes": [
            f"數據來源: {os.path.basename(edf_path)}",
            f"自動導入 {len(parsed_tasks)} 個任務，{len(sections)} 個階段",
            f"工期: {project_start} ~ {project_end}，共 {num_days} 日曆天",
        ],
    }
    if name_note:
        config["notes"].append("ℹ " + name_note)

    if output_json:
        with open(output_json, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        print(f"  已導出: {output_json}")

    # ---- 識別報告 ----
    report = {
        "source": os.path.basename(edf_path),
        "sheet": ws.title,
        "header_row": header_row,
        "columns": {field: {"index": idx + 1, "header": headers[idx]} for field, idx in mapping.items()},
        "metadata_found": meta,
        "project_name_resolved": project_name,
        "name_note": name_note,
        "task_count": len(parsed_tasks),
        "skipped_rows": skipped,
        "category_distribution": dict(sorted(cat_counts.items())),
        "tasks": [
            {"row": t["_excel_row"], "name": t["name"], "category": t["category"],
             "matched_keyword": t["matched_kw"], "deps": t["deps_raw"],
             "start": t["start_date"].isoformat(), "end": t["end_date"].isoformat(),
             "duration": t["duration"]}
            for t in parsed_tasks
        ],
        "unresolved_dependencies": unresolved,
        "fallback_dependencies": fallback_deps,
    }

    wb.close()

    if return_report:
        return config, report
    return config


# ================================================================
#  CLI 入口
# ================================================================

def main():
    if len(sys.argv) < 2:
        print(__doc__)
        print("用法: python edf_importer.py <EDF.xlsx> [-o output.json] [--project 名] "
              "[--subtitle 副標] [--sheet 表] [--report] [--profile 名]")
        sys.exit(1)

    edf_path = sys.argv[1]
    output_json = None
    project_name = ""
    subtitle = ""
    sheet_name = None
    do_report = False
    profile = None

    i = 2
    while i < len(sys.argv):
        a = sys.argv[i]
        if a == "-o" and i + 1 < len(sys.argv):
            output_json = sys.argv[i + 1]; i += 2
        elif a == "--project" and i + 1 < len(sys.argv):
            project_name = sys.argv[i + 1]; i += 2
        elif a == "--subtitle" and i + 1 < len(sys.argv):
            subtitle = sys.argv[i + 1]; i += 2
        elif a == "--sheet" and i + 1 < len(sys.argv):
            sheet_name = sys.argv[i + 1]; i += 2
        elif a == "--profile" and i + 1 < len(sys.argv):
            profile = sys.argv[i + 1]; i += 2
        elif a == "--report":
            do_report = True; i += 1
        else:
            i += 1

    print(f"=== EDF 自適應導入工具 v3.0 ===")
    print(f"來源: {edf_path}")

    if do_report:
        config, report = import_edf(
            edf_path, project_name=project_name, subtitle=subtitle,
            output_json=output_json, sheet_name=sheet_name, profile=profile,
            return_report=True,
        )
        # 輸出識別報告
        rep_path = os.path.splitext(edf_path)[0] + "_recognition_report.json"
        with open(rep_path, "w", encoding="utf-8") as f:
            json.dump(report, f, ensure_ascii=False, indent=2)
        _print_report(report)
        print(f"\n  識別報告已導出: {rep_path}")
    else:
        config = import_edf(
            edf_path, project_name=project_name, subtitle=subtitle,
            output_json=output_json, sheet_name=sheet_name, profile=profile,
        )

    print(f"\n✅ 導入完成")
    print(f"   項目: {config['project']['name']}")
    print(f"   任務: {sum(len(s['tasks']) for s in config['sections'])} 個")
    print(f"   階段: {len(config['sections'])} 個")
    print(f"   工期: {config['project']['num_days']} 日曆天")

    if output_json is None and not do_report:
        out = os.path.splitext(edf_path)[0] + "_gantt_config.json"
        with open(out, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        print(f"   已自動導出: {out}")


def _print_report(report: dict):
    print("\n" + "=" * 60)
    print("  識別報告 (表頭 / 列頭 / 內容 / 關鍵詞)")
    print("=" * 60)
    print(f"  來源: {report['source']}  工作表: {report['sheet']}")
    print(f"  表頭行: 第 {report['header_row']} 行")
    print(f"  列映射:")
    for field, info in report["columns"].items():
        print(f"    - {field:<12} ← 第{info['index']}列 '{info['header']}'")
    print(f"  元數據:")
    for k, v in report["metadata_found"].items():
        print(f"    - {k:<14}: {v}")
    print(f"  項目名稱(解析): {report['project_name_resolved']}")
    if report["name_note"]:
        print(f"  ℹ {report['name_note']}")
    print(f"  任務數: {report['task_count']}  跳過行: {len(report['skipped_rows'])}")
    print(f"  分類分佈: {report['category_distribution']}")
    if report["unresolved_dependencies"]:
        print(f"  ⚠ 未解析依賴: {report['unresolved_dependencies']}")
    if report.get("fallback_dependencies"):
        print(f"  ↳ 依賴回退(指向被跳過行→最近行): {report['fallback_dependencies']}")
    print("=" * 60)


if __name__ == "__main__":
    main()
