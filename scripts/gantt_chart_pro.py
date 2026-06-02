#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GanttChart Pro v14.9.1 — 修复闪退 + 摘要加深 + 恢复格线
"""
import re, json, copy, sys, traceback
from pathlib import Path
from datetime import datetime, timedelta
from typing import List, Tuple, Dict, Optional, Any

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel
from openpyxl.cell.cell import MergedCell


# ---------- 工具函数：颜色变暗 ----------
def darken_hex(hex_color, factor=0.7):
    """将十六进制颜色变暗，factor 越小越暗（0~1）"""
    hex_color = hex_color.lstrip('#')
    if len(hex_color) < 6:
        hex_color = hex_color.zfill(6)
    r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
    r, g, b = int(r * factor), int(g * factor), int(b * factor)
    return f"{r:02X}{g:02X}{b:02X}"


# ---------- 配置加载 ----------
def load_config(config_path="config_v2.json"):
    default = {
        "colors": {
            "HEADER_BG": "1E3B4A", "HEADER_FG": "FFFFFF",
            "INFO_BG": "F0F4FA", "TASK_BG_END": "4A7A9C",
            "MILESTONE_BG": "D9893E", "DEPENDENCY_MARKER": "C0503D",
            "WEEKEND_BG": "F7F9FC", "DARK_WEEKEND_BG": "9EABBA",
            "ALTERNATE_ODD": "F4F7FB", "ALTERNATE_EVEN": "FFFFFF",
            "GRID": "C5D1DF", "GRID_LIGHT": "DDE3EB", "GRID_BOLD": "1E3B4A",
            "BAR_BORDER": "B3C0D0", "FONT_COLOR": "2C3E50",
            "TODAY_COL": "FCEFE3", "TODAY_LINE": "D9534F",
            "SUMMARY_BAR": "2C3E50"
        },
        "stage_colors": {},
        "print": {"orientation": "landscape", "fit_to_pages_wide": 1, "fit_to_pages_tall": 0},
        "gantt_layout": {
            "right_panel_width_chars": 180.0,
            "preferred_interval_days": [1,2,3,4,5,7,10,14,21,30],
            "week_grouping": False,
            "week_start": 6
        },
        "calendar": {
            "type": "ask_user",
            "options": ["calendar_days", "workdays"],
            "option_labels": {
                "calendar_days": "日曆天（包含周末与节假日）",
                "workdays": "工作天（週一至週五，排除假期，周末藍灰）"
            },
            "holidays": []
        },
        "schemes": {}
    }
    config = copy.deepcopy(default)
    if Path(config_path).exists():
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                user = json.load(f)
            _deep_update(config, user)
            print(f"✓ 已加载配置: {config_path}")
        except Exception as e:
            print(f"⚠️ 加载配置失败: {e}，使用默认值")
    return config

def load_style(style_path="gantt_styles.json"):
    default = {
        "column_widths": {"A": 5.0, "B": 30.0, "C": 11.0, "D": 6.0, "E": 11.0},
        "row_heights": {
            "title_row": 36, "info_row": 20, "header_row": 28,
            "task_row": None, "scale_year": 20, "scale_month": 20,
            "scale_week": 18, "scale_day": 18, "scale_weekday": 18
        },
        "fonts": {
            "default_name": "微軟正黑體",
            "sizes": {"title": 16, "header": 11, "task": 11, "scale": 9}
        },
        "borders": {
            "grid_color_ref": "GRID", "grid_color_bold_ref": "GRID_BOLD",
            "regular_border_style": "thin", "bold_border_style": "medium",
            "today_line_color_ref": "TODAY_LINE", "today_line_style": "medium"
        },
        "misc": {"freeze_panes": "F7", "print_title_rows": "1:6"}
    }
    style = copy.deepcopy(default)
    if Path(style_path).exists():
        try:
            with open(style_path, 'r', encoding='utf-8') as f:
                user = json.load(f)
            _deep_update(style, user)
            print(f"✓ 已加载样式: {style_path}")
        except Exception as e:
            print(f"⚠️ 加载样式失败: {e}，使用默认值")
    return style

def _deep_update(target, source):
    for key, value in source.items():
        if key in target and isinstance(target[key], dict) and isinstance(value, dict):
            _deep_update(target[key], value)
        else:
            target[key] = value


# ---------- 日期处理 ----------
def parse_date(val):
    if val is None: return None
    if isinstance(val, datetime): return val.replace(hour=0, minute=0, second=0)
    if isinstance(val, (int, float)):
        try: return from_excel(val).replace(hour=0, minute=0, second=0)
        except: pass
    s = str(val).strip()
    if not s: return None

    # 常见日期格式
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%d-%m-%Y"):
        try: return datetime.strptime(s, fmt)
        except: continue

    # 中文日期格式（如 2026年5月25日星期一 或 2026年5月25日）
    import re
    m = re.match(r'(\d{4})年(\d{1,2})月(\d{1,2})日', s)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return datetime(y, mo, d)

    return None

def generate_full_date_sequence(start, end):
    dates = []
    cur = start.replace(hour=0, minute=0, second=0, microsecond=0)
    end_date = end.replace(hour=0, minute=0, second=0, microsecond=0)
    while cur <= end_date:
        dates.append(cur)
        cur += timedelta(days=1)
    return dates

def get_chinese_weekday(dt): return ['一', '二', '三', '四', '五', '六', '日'][dt.weekday()]


# ---------- 时间刻度（非周模式） ----------
def build_time_scales(ws, start_col, segments, style, colors, start_row=2, non_workdays=None):
    if non_workdays is None: non_workdays = set()
    row_year = start_row
    row_month = start_row + 1
    row_range = start_row + 2
    row_weekday = start_row + 3

    for row, key in [(row_year, "scale_year"), (row_month, "scale_month"),
                     (row_range, "scale_day"), (row_weekday, "scale_weekday")]:
        ws.row_dimensions[row].height = style["row_heights"].get(key, 20)

    font_scale = Font(name=style["fonts"]["default_name"], size=style["fonts"]["sizes"].get("scale", 9))
    border_cfg = style["borders"]
    grid_light = colors.get("GRID_LIGHT", "DDE3EB")
    bold_color = colors.get(border_cfg["grid_color_bold_ref"], "1E3B4A")
    thin_side = Side(style=border_cfg["regular_border_style"], color=grid_light)
    bold_side = Side(style=border_cfg["bold_border_style"], color=bold_color)
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    bg_fill = PatternFill(start_color=colors.get("INFO_BG", "F0F4FA"),
                          end_color=colors.get("INFO_BG", "F0F4FA"), fill_type="solid")
    dark_weekend = colors.get("DARK_WEEKEND_BG", "9EABBA")
    non_work_fill = PatternFill(start_color=dark_weekend, end_color=dark_weekend, fill_type="solid")

    # 年份
    year_groups = []
    cur_year, yr_start = None, 0
    for i, (seg_start, _) in enumerate(segments):
        y = seg_start.year
        if y != cur_year:
            if cur_year is not None: year_groups.append((yr_start, i-1, cur_year))
            cur_year = y; yr_start = i
    if cur_year is not None: year_groups.append((yr_start, len(segments)-1, cur_year))
    for s, e, y in year_groups:
        if s != e: ws.merge_cells(start_row=row_year, start_column=start_col+s, end_row=row_year, end_column=start_col+e)
        cell = ws.cell(row_year, start_col+s, value=f"{y}年")
        cell.font = font_scale; cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = bg_fill; cell.border = thin_border

    # 月份
    month_groups = []
    cur_month, mo_start = None, 0
    for i, (seg_start, _) in enumerate(segments):
        mk = (seg_start.year, seg_start.month)
        if mk != cur_month:
            if cur_month is not None: month_groups.append((mo_start, i-1, cur_month))
            cur_month = mk; mo_start = i
    if cur_month is not None: month_groups.append((mo_start, len(segments)-1, cur_month))
    for s, e, (y, m) in month_groups:
        if s != e: ws.merge_cells(start_row=row_month, start_column=start_col+s, end_row=row_month, end_column=start_col+e)
        cell = ws.cell(row_month, start_col+s, value=f"{m}月")
        cell.font = font_scale; cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = bg_fill; cell.border = thin_border

    # 日期段和星期
    for i, (seg_start, seg_end) in enumerate(segments):
        col = start_col + i
        all_non = all((seg_start + timedelta(days=d)).date() in non_workdays for d in range((seg_end - seg_start).days + 1))
        fill_upper = non_work_fill if all_non else bg_fill
        cell_day = ws.cell(row_range, col, value=seg_start.strftime("%d"))
        cell_day.font = font_scale
        cell_day.alignment = Alignment(horizontal='center', vertical='center')
        cell_day.fill = fill_upper
        cell_day.border = thin_border

        wd = get_chinese_weekday(seg_start)
        cell_wd = ws.cell(row_weekday, col, value=wd)
        cell_wd.font = Font(name=style["fonts"]["default_name"], size=style["fonts"]["sizes"].get("scale", 9), color="000000")
        cell_wd.alignment = Alignment(horizontal='center', vertical='center')
        cell_wd.fill = bg_fill
        cell_wd.border = thin_border

    # 月份分组加粗竖线
    for s, e, _ in month_groups:
        for r in [row_year, row_month, row_range, row_weekday]:
            cell = ws.cell(r, start_col + e)
            cell.border = Border(left=cell.border.left if cell.border else thin_side,
                                 right=bold_side,
                                 top=cell.border.top if cell.border else thin_side,
                                 bottom=cell.border.bottom if cell.border else thin_side)

    return row_weekday


# ---------- 时间刻度（周模式） ----------
def build_week_scales(ws, start_col, all_dates, style, colors, week_start=6, start_row=2, non_workdays=None):
    if non_workdays is None: non_workdays = set()
    border_cfg = style["borders"]
    grid_light = colors.get("GRID_LIGHT", "DDE3EB")
    bold_color = colors.get(border_cfg["grid_color_bold_ref"], "1E3B4A")
    thin_side = Side(style=border_cfg["regular_border_style"], color=grid_light)
    bold_side = Side(style=border_cfg["bold_border_style"], color=bold_color)
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    bg_fill = PatternFill(start_color=colors.get("INFO_BG", "F0F4FA"),
                          end_color=colors.get("INFO_BG", "F0F4FA"), fill_type="solid")
    dark_weekend = colors.get("DARK_WEEKEND_BG", "9EABBA")
    non_work_fill = PatternFill(start_color=dark_weekend, end_color=dark_weekend, fill_type="solid")
    font_label = Font(name=style["fonts"]["default_name"], size=style["fonts"]["sizes"].get("scale", 9))
    font_wd = Font(name=style["fonts"]["default_name"], size=style["fonts"]["sizes"].get("scale", 9), color="000000")

    row_year = start_row
    row_month = start_row + 1
    row_week = start_row + 2
    row_wd = start_row + 3

    ws.row_dimensions[row_year].height = style["row_heights"]["scale_year"]
    ws.row_dimensions[row_month].height = style["row_heights"]["scale_month"]
    ws.row_dimensions[row_week].height = style["row_heights"].get("scale_week", 18)
    ws.row_dimensions[row_wd].height = style["row_heights"]["scale_weekday"]

    # 按周分组
    weeks, week = [], []
    for d in all_dates:
        if not week:
            week.append(d)
        else:
            if d.weekday() == week_start:
                weeks.append(week)
                week = [d]
            else:
                week.append(d)
    if week: weeks.append(week)

    col_date_map = []
    col = start_col
    for wds in weeks:
        for day in wds:
            col_date_map.append((col, day))
            col += 1

    # 年份合并
    year_groups = []
    cur_year, yr_start = None, 0
    for idx, (c, d) in enumerate(col_date_map):
        y = d.year
        if y != cur_year:
            if cur_year is not None: year_groups.append((yr_start, idx-1, cur_year))
            cur_year = y; yr_start = idx
    if cur_year is not None: year_groups.append((yr_start, len(col_date_map)-1, cur_year))
    for s, e, y in year_groups:
        if s != e: ws.merge_cells(start_row=row_year, start_column=col_date_map[s][0], end_row=row_year, end_column=col_date_map[e][0])
        cell = ws.cell(row_year, col_date_map[s][0], value=f"{y}年")
        cell.font = font_label; cell.alignment = Alignment(horizontal='center', vertical='center'); cell.fill = bg_fill; cell.border = thin_border

    # 月份合并
    month_groups = []
    cur_month, mo_start = None, 0
    for idx, (c, d) in enumerate(col_date_map):
        mk = (d.year, d.month)
        if mk != cur_month:
            if cur_month is not None: month_groups.append((mo_start, idx-1, cur_month))
            cur_month = mk; mo_start = idx
    if cur_month is not None: month_groups.append((mo_start, len(col_date_map)-1, cur_month))
    for s, e, (y, m) in month_groups:
        if s != e: ws.merge_cells(start_row=row_month, start_column=col_date_map[s][0], end_row=row_month, end_column=col_date_map[e][0])
        cell = ws.cell(row_month, col_date_map[s][0], value=f"{m}月")
        cell.font = font_label; cell.alignment = Alignment(horizontal='center', vertical='center'); cell.fill = bg_fill; cell.border = thin_border

    # 周标签和星期
    col = start_col
    for wds in weeks:
        wstart, wend = wds[0], wds[-1]
        if len(wds) > 1:
            ws.merge_cells(start_row=row_week, start_column=col, end_row=row_week, end_column=col+len(wds)-1)
        cell_week = ws.cell(row_week, col, value=f"{wstart.month}/{wstart.day} - {wend.month}/{wend.day}")
        cell_week.font = font_label
        cell_week.alignment = Alignment(horizontal='center', vertical='center')
        cell_week.fill = bg_fill

        for i, day in enumerate(wds):
            c = col + i
            ws.cell(row_week, c).fill = bg_fill
            ws.cell(row_week, c).border = thin_border
            day_cell = ws.cell(row_wd, c, value=get_chinese_weekday(day))
            day_cell.font = font_wd
            day_cell.alignment = Alignment(horizontal='center', vertical='center')
            day_cell.fill = bg_fill
            if day.date() in non_workdays and non_workdays:
                ws.cell(row_year, c).fill = non_work_fill
                ws.cell(row_month, c).fill = non_work_fill
                ws.cell(row_week, c).fill = non_work_fill
            day_cell.border = thin_border

        last_c = col + len(wds) - 1
        for r in (row_year, row_month, row_week, row_wd):
            cell = ws.cell(r, last_c)
            cell.border = Border(left=cell.border.left if cell.border else thin_side,
                                 right=bold_side,
                                 top=cell.border.top if cell.border else thin_side,
                                 bottom=cell.border.bottom if cell.border else thin_side)
        col += len(wds)

    return row_wd


# ---------- EDF 解析 ----------
def extract_metadata(ws):
    meta = {"project_name": "", "project_no": "", "constructor": "", "prepare_date": "", "total_period": ""}
    for row in range(1, 15):
        for col in range(ws.max_column - 8, ws.max_column + 1):
            if col < 1: continue
            val = str(ws.cell(row, col).value or "")
            if not val: continue
            right = ws.cell(row, col+1).value
            right_str = str(right).strip() if right else ""
            if "工程項目名稱" in val: meta["project_name"] = right_str
            elif "工程項目編號" in val: meta["project_no"] = right_str
            elif "承建商" in val: meta["constructor"] = right_str
            elif "編制時間" in val:
                dt = parse_date(right)
                meta["prepare_date"] = dt.strftime("%Y-%m-%d") if dt else str(right)
            elif "總工期" in val: meta["total_period"] = right_str
    if not meta["project_name"]: meta["project_name"] = "项目"
    if not meta["prepare_date"]: meta["prepare_date"] = datetime.now().strftime("%Y-%m-%d")
    return meta

def get_task_stage(seq):
    m = re.match(r'^([A-Z])(?:\d|\.)', seq)
    return m.group(1) if m else None

def parse_tasks(ws):
    header_row, col_map = None, {}
    for row_num in range(1, 15):
        vals = [str(c.value or "") for c in ws[row_num]]
        # 检测表头：同时包含“序號”和“施工內容”等
        has_seq = any(any(kw in v for kw in ("序號","項次","No","#")) for v in vals)
        has_name = any(any(kw in v for kw in ("施工內容","任务名称","任務名稱","項目名稱","Description")) for v in vals)
        if has_seq and has_name:
            header_row = row_num
            for i, v in enumerate(vals, 1):
                if any(kw in v for kw in ("序號","項次","No","#")):
                    col_map["seq"] = i
                elif any(kw in v for kw in ("施工內容","任务名称","任務名稱","項目名稱","Description")):
                    col_map["name"] = i
                elif any(kw in v for kw in ("開始日期","开始时间","開始時間","開工","Start","Sd")):
                    col_map["start"] = i
                elif any(kw in v for kw in ("完成日期","完成时间","完成時間","完工","End","Ed","結束")):
                    col_map["end"] = i
            break
    if not header_row: raise ValueError("找不到任务表头")
    tasks = []
    for row in ws.iter_rows(min_row=header_row+1):
        seq = row[col_map["seq"]-1].value if "seq" in col_map else None
        name = row[col_map["name"]-1].value if "name" in col_map else None
        if not seq or not name: continue
        is_cat = bool(re.match(r'^[A-Z]\.', str(seq)))
        start = parse_date(row[col_map["start"]-1].value) if "start" in col_map else None
        end = parse_date(row[col_map["end"]-1].value) if "end" in col_map else None
        if start and end and start > end: start, end = end, start
        tasks.append({"seq": str(seq).strip(), "name": str(name).strip(),
                      "start": start, "end": end, "is_category": is_cat,
                      "stage": get_task_stage(str(seq)) if not is_cat else None})
    return tasks


# ---------- 辅助 ----------
def apply_table_border(ws, sr, er, sc, ec, style, colors):
    border_cfg = style["borders"]
    grid_color = colors.get("GRID_LIGHT", "DDE3EB")
    bold_color = colors.get(border_cfg["grid_color_bold_ref"], "1E3B4A")
    thin = Side(style=border_cfg["regular_border_style"], color=grid_color)
    bold = Side(style=border_cfg["bold_border_style"], color=bold_color)
    for r in range(sr, er+1):
        for c in range(sc, ec+1):
            cell = ws.cell(r, c)
            if isinstance(cell, MergedCell):
                continue
            cell.border = Border(left=bold if c==sc else thin,
                                 right=bold if c==ec else thin,
                                 top=bold if r==sr else thin,
                                 bottom=bold if r==er else thin)

def auto_fit_left_columns(ws, tasks, style):
    max_len = max((len(t["name"]) for t in tasks if t["name"]), default=20)
    ws.column_dimensions['A'].width = style["column_widths"]["A"]
    ws.column_dimensions['B'].width = max(style["column_widths"]["B"], max_len*1.2)
    ws.column_dimensions['C'].width = 11.0
    ws.column_dimensions['D'].width = 6.0
    ws.column_dimensions['E'].width = 11.0


# ---------- 主流程 ----------
def create_gantt_pro(input_path, output_path=None):
    config = load_config()
    style = load_style()

    # 配色方案
    schemes = config.get("schemes", {})
    active_scheme = config.get("active_scheme")
    if schemes:
        print("\n可用的配色方案：")
        slist = list(schemes.keys())
        for i, name in enumerate(slist, 1):
            print(f"  {i}. {schemes[name].get('name', name)} ({name})")
        print(f"  0. 使用默认 ({active_scheme})")
        choice = input("请选择配色方案: ").strip()
        if choice.isdigit():
            n = int(choice)
            if 1 <= n <= len(slist): active_scheme = slist[n-1]
    if active_scheme and active_scheme in schemes:
        s = schemes[active_scheme]
        colors = s.get("colors", config["colors"])
        stage_colors = s.get("stage_colors", config.get("stage_colors", {}))
    else:
        colors = config["colors"]
        stage_colors = config.get("stage_colors", {})

    # 读取文件
    print(f"\n📂 读取文件: {input_path}")
    try:
        wb = openpyxl.load_workbook(input_path, data_only=True)
        ws_src = wb.active
    except Exception as e:
        print(f"❌ 文件打开失败: {e}")
        return False

    metadata = extract_metadata(ws_src)
    tasks = parse_tasks(ws_src)
    valid = [t for t in tasks if t["start"] and t["end"] and not t["is_category"]]
    if not valid:
        print("❌ 无有效任务")
        return False

    start_date = min(t["start"] for t in valid)
    end_date = max(t["end"] for t in valid)
    print(f"📅 工期: {start_date.strftime('%Y-%m-%d')} ～ {end_date.strftime('%Y-%m-%d')}")

    # 日历类型
    cal = config["calendar"]
    if cal.get("type") == "ask_user":
        print("\n请选择日历类型：")
        opts = cal.get("options", ["calendar_days", "workdays"])
        labels = cal.get("option_labels", {})
        for i, o in enumerate(opts, 1): print(f"  {i}. {labels.get(o, o)}")
        use_workdays = input("请输入数字 (1/2)，直接回车为日曆天: ").strip() == '2'
    else:
        use_workdays = (cal.get("type") == "workdays")

    all_dates = generate_full_date_sequence(start_date, end_date)
    holidays_list = cal.get("holidays", [])
    holiday_set = set()
    for h in holidays_list:
        try: holiday_set.add(datetime.strptime(h, "%Y-%m-%d").date())
        except: pass

    if use_workdays:
        non_workdays = set()
        for d in all_dates:
            if d.weekday() >= 5 or d.date() in holiday_set:
                non_workdays.add(d.date())
    else:
        non_workdays = set()

    print(f"📅 日历类型: {'工作日' if use_workdays else '日曆天'}, 总天数: {len(all_dates)} 天")

    # 时间显示模式
    print("\n请选择时间显示模式：")
    print("  1. 按周分组（每天一列，每周一栏）")
    print("  2. 合并多天一列（压缩显示）")
    mode_choice = input("请输入数字 (1/2)，直接回车默认为合并: ").strip()
    week_grouping = (mode_choice == '1')

    layout = config["gantt_layout"]
    week_start = layout.get("week_start", 6)

    if week_grouping:
        interval_days = 1
        segments = [(d, d) for d in all_dates]
        actual_columns = len(all_dates)
        right_width = layout.get("right_panel_width_chars", 180.0)
        col_width = max(2.0, right_width / actual_columns)
        print(f"📊 按周分组模式: {actual_columns} 列, 列宽 {col_width:.2f}")
    else:
        right_width = layout.get("right_panel_width_chars", 180.0)
        preferred = layout.get("preferred_interval_days", [1,2,3,4,5,7,10,14,21,30])
        def suggest(days):
            for interval in preferred:
                cols = (days + interval - 1) // interval
                cw = right_width / cols
                if 2.0 <= cw <= 15.0: return interval, cols, cw
            best, best_diff = preferred[-1], float('inf')
            for interval in preferred:
                cols = (days + interval - 1) // interval
                cw = right_width / cols
                if abs(cw-6.0) < best_diff: best_diff = abs(cw-6.0); best = interval
            cols = (days + best - 1) // best
            return best, cols, right_width/cols
        rec_int, rec_cols, rec_w = suggest(len(all_dates))
        print(f"\n请选择时间颗粒度：\n  0. 自动推荐（{rec_int} 天一列，约 {rec_cols} 列）")
        for i, d in enumerate(preferred, 1):
            print(f"  {i}. {d} 天一列（约 {(len(all_dates)+d-1)//d} 列）")
        print(f"  {len(preferred)+1}. 手动输入")
        choice = input("请输入选项数字: ").strip()
        if choice == '0': interval_days = rec_int
        elif choice == str(len(preferred)+1):
            try: interval_days = max(1, min(99, int(input("间隔天数（1-99）: "))))
            except: interval_days = rec_int
        else:
            try:
                idx = int(choice)-1
                interval_days = preferred[idx] if 0 <= idx < len(preferred) else rec_int
            except: interval_days = rec_int
        segments = [(all_dates[i], all_dates[min(i+interval_days-1, len(all_dates)-1)])
                    for i in range(0, len(all_dates), interval_days)]
        actual_columns = len(segments)
        col_width = max(2.0, right_width / actual_columns)
        print(f"📊 颗粒度: {interval_days} 天/列 → {actual_columns} 列，列宽 {col_width:.2f}")

    # 创建工作表
    ws = wb.create_sheet("甘特图")
    for cl, w in style["column_widths"].items():
        ws.column_dimensions[cl].width = w
    start_col = 6
    for i in range(actual_columns):
        ws.column_dimensions[get_column_letter(start_col + i)].width = col_width

    # 时间刻度
    scale_start_row = 2
    if week_grouping:
        last_scale_row = build_week_scales(ws, start_col, all_dates, style, colors, week_start, scale_start_row, non_workdays)
    else:
        last_scale_row = build_time_scales(ws, start_col, segments, style, colors, scale_start_row, non_workdays)

    # 信息区（行2-5，A-E列）
    info_bg = PatternFill(start_color=colors["INFO_BG"], end_color=colors["INFO_BG"], fill_type="solid")
    font_color = colors.get("FONT_COLOR", "2C3E50")
    info_font = Font(size=style["fonts"]["sizes"]["task"], color=font_color)
    ws.merge_cells('A2:B2')
    c2 = ws.cell(2, 1, value=f"工程编号：{metadata['project_no']}")
    c2.font = info_font; c2.alignment = Alignment(horizontal='left', vertical='center'); c2.fill = info_bg
    for col in range(3, 6): ws.cell(2, col).fill = info_bg
    ws.merge_cells('A3:E3')
    c3 = ws.cell(3, 1, value=f"工程施工期：{metadata['total_period']}")
    c3.font = info_font; c3.alignment = Alignment(horizontal='left', vertical='center'); c3.fill = info_bg
    ws.merge_cells('A4:C4')
    c4 = ws.cell(4, 1, value=f"承建商：{metadata['constructor']}")
    c4.font = info_font; c4.alignment = Alignment(horizontal='left', vertical='center'); c4.fill = info_bg
    for col in range(4, 6): ws.cell(4, col).fill = info_bg
    ws.merge_cells('A5:B5')
    c5 = ws.cell(5, 1, value=f"编制日期：{metadata['prepare_date']}")
    c5.font = info_font; c5.alignment = Alignment(horizontal='left', vertical='center'); c5.fill = info_bg
    for col in range(3, 6): ws.cell(5, col).fill = info_bg

    for r in range(2, 6):
        if ws.row_dimensions[r].height is None:
            ws.row_dimensions[r].height = style["row_heights"]["info_row"]

    # 标题行
    total_cols = start_col + actual_columns - 1
    ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
    tc = ws.cell(1, 1, value=f"{metadata['project_name']} - 甘特图")
    tc.font = Font(name=style["fonts"]["default_name"], size=style["fonts"]["sizes"]["title"], bold=True, color=colors.get("HEADER_FG", "FFFFFF"))
    tc.alignment = Alignment(horizontal='center', vertical='center')
    tc.fill = PatternFill(start_color=colors["HEADER_BG"], end_color=colors["HEADER_BG"], fill_type="solid")
    ws.row_dimensions[1].height = style["row_heights"]["title_row"]

    # 表头
    header_row = 6
    hdrs = ["序號", "施工內容", "開始", "天數", "完成"]
    aligns = ["center", "left", "center", "center", "center"]
    hfill = PatternFill(start_color=colors["HEADER_BG"], end_color=colors["HEADER_BG"], fill_type="solid")
    hfont = Font(bold=True, color=colors.get("HEADER_FG", "FFFFFF"), size=style["fonts"]["sizes"]["header"])
    for i, (h, al) in enumerate(zip(hdrs, aligns), 1):
        cell = ws.cell(header_row, i, value=h)
        cell.font = hfont; cell.fill = hfill; cell.alignment = Alignment(horizontal=al, vertical='center')
    ws.row_dimensions[header_row].height = style["row_heights"]["header_row"]

    # 边框样式
    border_cfg = style["borders"]
    grid_light = colors.get("GRID_LIGHT", "DDE3EB")
    thin_side = Side(style=border_cfg["regular_border_style"], color=grid_light)
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    dark_weekend = colors.get("DARK_WEEKEND_BG", "9EABBA")
    non_work_fill = PatternFill(start_color=dark_weekend, end_color=dark_weekend, fill_type="solid")

    odd_bg = colors.get("ALTERNATE_ODD", "F4F7FB")
    even_bg = colors.get("ALTERNATE_EVEN", "FFFFFF")
    odd_fill = PatternFill(start_color=odd_bg, end_color=odd_bg, fill_type="solid")
    even_fill = PatternFill(start_color=even_bg, end_color=even_bg, fill_type="solid")

    task_font = Font(name=style["fonts"]["default_name"], size=style["fonts"]["sizes"]["task"])
    # 子任务横道边框（略微可见的轮廓）
    bar_border_normal = Border(left=thin_side, right=thin_side,
                               top=thin_side, bottom=thin_side,
                               outline=Side(style="thin", color=colors.get("BAR_BORDER", "B3C0D0")))

    # 分类汇总时间
    cat_tasks = {t["seq"]: t for t in tasks if t["is_category"]}
    for cat_seq, cat in cat_tasks.items():
        prefix = cat_seq.rstrip('.')
        children = [t for t in tasks if t["seq"].startswith(prefix) and not t["is_category"] and t["start"] and t["end"]]
        if children:
            cat["start"] = min(c["start"] for c in children)
            cat["end"] = max(c["end"] for c in children)

    # 任务数据
    task_start_row = header_row + 1
    for idx, task in enumerate(tasks):
        row_num = task_start_row + idx
        row_fill = odd_fill if idx % 2 == 0 else even_fill
        for c in range(1, total_cols+1):
            cell = ws.cell(row_num, c)
            if isinstance(cell, MergedCell): continue
            cell.fill = row_fill

        ws.cell(row_num, 1, value=task["seq"]).font = task_font
        ws.cell(row_num, 1).alignment = Alignment(horizontal='center', vertical='center')
        nc = ws.cell(row_num, 2, value=task["name"])
        nc.font = Font(name=style["fonts"]["default_name"], size=style["fonts"]["sizes"]["task"],
                       bold=task["is_category"],
                       color=colors.get("FONT_COLOR", "2C3E50") if task["is_category"] else "000000")
        nc.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        if task["start"]:
            c3 = ws.cell(row_num, 3, value=task["start"])
            c3.number_format = "yyyy/mm/dd"; c3.font = task_font; c3.alignment = Alignment(horizontal='center', vertical='center')
        if task["start"] and task["end"]:
            dur = (task["end"] - task["start"]).days + 1
            c4 = ws.cell(row_num, 4, value=dur)
            c4.font = task_font; c4.alignment = Alignment(horizontal='center', vertical='center')
        if task["end"]:
            c5 = ws.cell(row_num, 5, value=task["end"])
            c5.number_format = "yyyy/mm/dd"; c5.font = task_font; c5.alignment = Alignment(horizontal='center', vertical='center')

        # 周末列填充
        if non_workdays:
            for si, (seg_start, seg_end) in enumerate(segments):
                col = start_col + si
                all_non = True
                d = seg_start
                while d <= seg_end:
                    if d.date() not in non_workdays:
                        all_non = False
                        break
                    d += timedelta(days=1)
                if all_non:
                    cell = ws.cell(row_num, col)
                    if isinstance(cell, MergedCell): continue
                    cell.fill = non_work_fill

        # 横道（逐格绘制，不合并）
        if task["start"] and task["end"]:
            if task["is_category"]:
                # 摘要任务：基于同阶段颜色加深
                base_color = stage_colors.get(task.get("stage"), colors.get("TASK_BG_END", "4A7A9C"))
                bar_color = darken_hex(base_color, 0.7)
                bar_fill = PatternFill(start_color=bar_color, end_color=bar_color, fill_type="solid")
                bar_border = Border(left=thin_side, right=thin_side,
                                    top=Side(style="medium", color=colors.get("GRID_BOLD", "1E3B4A")),
                                    bottom=Side(style="medium", color=colors.get("GRID_BOLD", "1E3B4A")))
            else:
                stage = task.get("stage")
                bc = stage_colors.get(stage, colors.get("TASK_BG_END", "4A7A9C"))
                bar_fill = PatternFill(start_color=bc, end_color=bc, fill_type="solid")
                bar_border = bar_border_normal

            bar_cols = []
            for si, (seg_s, seg_e) in enumerate(segments):
                if non_workdays:
                    if all((seg_s + timedelta(days=d)).date() in non_workdays for d in range((seg_e - seg_s).days + 1)):
                        continue
                col = start_col + si
                if max(task["start"].date(), seg_s.date()) <= min(task["end"].date(), seg_e.date()):
                    bar_cols.append(col)

            for col in bar_cols:
                cell = ws.cell(row_num, col)
                cell.fill = bar_fill
                cell.border = bar_border
                if task["is_category"]:
                    cell.value = " "
                    cell.font = Font(color="FFFFFF", bold=True, size=9)
                else:
                    if dur == 1 and "里程碑" in task["name"]:
                        cell.value = "◆"
                        cell.font = Font(color="FFFFFF", bold=True, size=8)

        # 补充整行细边框
        for c in range(1, total_cols+1):
            cell = ws.cell(row_num, c)
            if isinstance(cell, MergedCell): continue
            if cell.border is None:
                cell.border = thin_border

        if style["row_heights"].get("task_row") is None:
            ws.row_dimensions[row_num].height = None
        else:
            ws.row_dimensions[row_num].height = style["row_heights"]["task_row"]

    # 数据区竖线（极淡细线）
    for si in range(actual_columns):
        col = start_col + si
        for r in range(task_start_row, task_start_row + len(tasks)):
            cell = ws.cell(r, col)
            if isinstance(cell, MergedCell):
                continue
            cell.border = thin_border

    last_row = task_start_row + len(tasks) - 1
    apply_table_border(ws, 1, last_row, 1, total_cols, style, colors)
    auto_fit_left_columns(ws, tasks, style)

    # 打印设置
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = config["print"].get("orientation", "landscape")
    def cm2inch(cm): return cm/2.54
    ws.page_margins.left = cm2inch(1.5); ws.page_margins.right = cm2inch(1.0)
    ws.page_margins.top = cm2inch(2.0); ws.page_margins.bottom = cm2inch(1.5)
    ws.page_margins.header = cm2inch(1.3); ws.page_margins.footer = cm2inch(1.3)
    ws.page_setup.horizontalCentered = True; ws.page_setup.verticalCentered = True
    ws.page_setup.fitToWidth = config["print"].get("fit_to_pages_wide", 1)
    ws.page_setup.fitToHeight = config["print"].get("fit_to_pages_tall", 0)
    ws.print_title_rows = style["misc"].get("print_title_rows", "1:6")
    ws.freeze_panes = style["misc"].get("freeze_panes", "F7")

    if not output_path:
            # 1. 配色方案名称（使用 active_scheme，若为空则用 default）
            scheme_name = active_scheme if active_scheme else "default"
            # 2. 日历类型显示
            cal_display = "工作日" if use_workdays else "日曆天"
            # 3. 显示模式（周模式或 X日粒度）
            mode_display = "周模式" if week_grouping else f"{interval_days}日粒度"
            # 4. 工程名（来自元数据，并清理非法字符）
            proj_name = metadata.get('project_name', Path(input_path).stem)
            import re
            proj_name_clean = re.sub(r'[\\/*?:"<>|]', '', proj_name)
            # 5. 组合最终文件名
            output_path = f"{scheme_name}_{cal_display}_{mode_display}_{proj_name_clean}.xlsx"
    wb.save(output_path)
    print(f"\n✅ 生成成功: {output_path}")
    return True


def main():
    print("="*70)
    print("GanttChart Pro v14.9.1 — 修复版")
    print("="*70)
    input_file = "EDF.xlsx"
    if not Path(input_file).exists():
        input_file = input("请拖入或输入 EDF 文件路径: ").strip()
        if not input_file:
            print("未提供文件，退出。")
            input("按 Enter 键退出...")
            return
    try:
        success = create_gantt_pro(input_file)
        if not success:
            print("\n生成失败，请检查文件格式。")
    except Exception as e:
        print(f"\n❌ 发生错误: {e}")
        traceback.print_exc()
    print("\n按 Enter 键退出...")
    input()

if __name__ == "__main__":
    main()