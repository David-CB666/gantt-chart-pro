#!/usr/bin/env python3
"""
Gantt Chart Generator v3.0 — Unified Rendering Engine
=====================================================
Integrates with the project configuration system:
  - config_v2.json      → color schemes, fields, calendar, i18n, validation, output
  - gantt_styles.json   → column widths, row heights, fonts, borders
  - time_utils.py       → date calculations, holiday support, month groupings

Usage:
    python gen_gantt.py                    → 啟動 GUI 可視化面板（雙擊默認模式）
    python gen_gantt.py [tasks.json|EDF.xlsx] [output.xlsx] [--project "Name"]

    tasks.json   — task data file in standard format (default: gantt_config.json)
    EDF.xlsx     — Excel spreadsheet, auto-detected and imported via edf_importer
    output.xlsx  — output path (auto-named from project name if omitted)
    --project    — override project name (EDF import only)
    --subtitle   — override subtitle (EDF import only)

Task data format (JSON):
    {
      "project": { "name": "...", "subtitle": "...", "start_date": "YYYY-MM-DD", ... },
      "sections": [
        { "title": "...", "tasks": [
            { "id": 1, "name": "...", "duration": 3, "start_day": 0,
              "deps": [], "category": "A", "milestone": false, "progress": 0 }
        ]}
      ],
      "critical_path": [1, 2, 3],  // optional, auto-calculated if omitted
      "legend": [ {"color": "...", "label": "..."} ],
      "notes": ["..."]
    }
"""

import json
import sys
import os
import math
import shutil
from datetime import date, timedelta, datetime
from typing import Optional, Dict, List, Set, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Local modules (same directory)
from time_utils import TimeModule

try:
    sys.stdout.reconfigure(encoding='utf-8')
except Exception:
    pass


# ================================================================
#  Constants & Color Utilities
# ================================================================

WEEKDAY_CN = ["一", "二", "三", "四", "五", "六", "日"]


def hex2rgb(h: str) -> tuple:
    h = h.lstrip("#")
    return tuple(int(h[i:i + 2], 16) for i in (0, 2, 4))


def rgb2hex(r: int, g: int, b: int) -> str:
    return f"{max(0, min(255, r)):02X}{max(0, min(255, g)):02X}{max(0, min(255, b)):02X}"


def lighten(color: str, f=0.4) -> str:
    r, g, b = hex2rgb(color)
    return rgb2hex(
        int(r + (255 - r) * f),
        int(g + (255 - g) * f),
        int(b + (255 - b) * f),
    )


def darken(color: str, f=0.25) -> str:
    r, g, b = hex2rgb(color)
    return rgb2hex(int(r * (1 - f)), int(g * (1 - f)), int(b * (1 - f)))


def tr(translations: dict, key: str, fallback: str = "") -> str:
    return translations.get(key, fallback) if translations else fallback


# ================================================================
#  Data Validator
# ================================================================

def validate_tasks(tasks: list, rules: dict) -> list:
    """Validate task data against config rules. Returns list of warnings."""
    warnings = []
    max_tasks = rules.get("max_tasks", 200)
    min_dur = rules.get("min_duration", 1)
    max_dur = rules.get("max_duration", 90)

    if len(tasks) > max_tasks:
        warnings.append(f"Task count ({len(tasks)}) exceeds max ({max_tasks})")

    for t in tasks:
        dur = t.get("duration", 1)
        if dur < min_dur:
            warnings.append(f"Task {t['id']} duration ({dur}) below min ({min_dur})")
        if dur > max_dur:
            warnings.append(f"Task {t['id']} duration ({dur}) exceeds max ({max_dur})")

        for dep_id in t.get("deps", []):
            if not any(x["id"] == dep_id for x in tasks):
                warnings.append(f"Task {t['id']} references unknown dep {dep_id}")

    return warnings


# ================================================================
#  GanttChart Renderer
# ================================================================

class GanttChart:
    """
    Data-driven Gantt chart renderer.
    Integrates config_v2.json + gantt_styles.json + time_utils.TimeModule.
    """

    INFO_BASE = ["seq", "name", "duration", "start", "end", "predecessor"]
    INFO_LABELS_FALLBACK = {
        "seq": "序號", "name": "項目名稱", "duration": "工期",
        "start": "開始", "end": "結束", "predecessor": "前置任務",
        "material": "材料訂貨期", "remark": "備註", "category": "類別",
    }

    def __init__(self, task_config: dict, output_path: str,
                 sys_config: dict = None, styles_config: dict = None):
        self.task_cfg = task_config
        self.output = output_path
        self.sys_cfg = sys_config or {}
        self.sty_cfg = styles_config or {}

        self._load_scheme()
        self._load_styles()
        self._load_calendar_cfg()
        self._load_layout()
        self._load_granularity()
        self._load_i18n()
        self._load_output_cfg()
        self._load_validation()
        self._load_project()
        self._build_calendar()
        self._process_tasks()
        self._validate()
        self._compute_cp()

        self.wb = Workbook()
        self._render()

    # ---- Config loaders ----

    def _load_scheme(self):
        scheme_name = self.sys_cfg.get("active_scheme", "blue_pro")
        schemes = self.sys_cfg.get("schemes", {})
        active = schemes.get(scheme_name, {})
        self.colors = active.get("colors", self.sys_cfg.get("colors", {}))
        self.stage_colors = active.get(
            "stage_colors", self.sys_cfg.get("stage_colors", {})
        )
        self.scheme_name = scheme_name

    def _load_styles(self):
        s = self.sty_cfg
        self.font_name = s.get("fonts", {}).get("default_name", "Arial")
        self.font_sizes = s.get("fonts", {}).get("sizes", {
            "title": 14, "header": 12, "task": 11, "scale": 10,
        })
        self.row_heights = s.get("row_heights", {})
        self.col_widths_style = s.get("column_widths", {})
        self.freeze_cfg = s.get("misc", {}).get("freeze_panes", "F7")
        self.border_style = s.get("borders", {})

    def _load_calendar_cfg(self):
        cal = self.sys_cfg.get("calendar", {})
        self.cal_type = cal.get("type", "calendar_days")
        if self.cal_type == "ask_user":
            self.cal_type = "calendar_days"
        self.work_week = cal.get("work_week", [1, 2, 3, 4, 5])
        self.holidays = set()
        for h in cal.get("holidays", []):
            try:
                p = h.split("-")
                self.holidays.add(date(int(p[0]), int(p[1]), int(p[2])))
            except Exception:
                pass
        self.makeup_days = set()
        for m in cal.get("makeup_days", []):
            try:
                p = m.split("-")
                self.makeup_days.add(date(int(p[0]), int(p[1]), int(p[2])))
            except Exception:
                pass

    def _load_layout(self):
        lay = self.sys_cfg.get("gantt_layout", {})
        self.target_cols = lay.get("target_columns", 30)
        self.info_col_widths = lay.get("column_widths", {})
        self.date_fmt = lay.get("date_format", "MM/dd")
        self.month_fmt = lay.get("month_format", "MM月")
        self.merge_months = lay.get("merge_months", True)
        self.show_weeks = lay.get("show_week_numbers", False)

    def _load_granularity(self):
        """Load and auto-determine time granularity.

        Supported modes: auto / day / week_grouped / week / compressed / month
          - day          : 每日一列（月份頂部分組）
          - week_grouped : 每日一列 + 每周一栏（週分組，需求①）
          - week         : 每週一列（ISO 週 Mon–Sun）
          - compressed   : 合併多天一列（壓縮顯示，需求②），days_per_col 可手動或自動
          - month        : 每月一列
          - auto         : 按跨度日/週/月 自動切換

        Auto logic (auto 模式):
          - total_days <= 120 → day
          - 121..450        → week
          - > 450           → month
        Manual override via granularity.manual config key.
        """
        g_cfg = self.sys_cfg.get("granularity", {})
        self.gran_auto = g_cfg.get("auto", True)
        self.gran_manual = g_cfg.get("manual", "day")
        self.gran_thresholds = g_cfg.get("thresholds", {"day_max": 120, "week_max": 450})
        self.gran_col_w = g_cfg.get("column_width",
                                    {"day": 4.5, "week": 3.0, "month": 2.0,
                                     "week_grouped": 4.5, "compressed": 6.0})
        self.gran_label_fmt = g_cfg.get("label_format",
                                        {"day": "d", "week": "W{week}", "month": "m",
                                         "week_grouped": "d", "compressed": "range"})
        # compressed 模式參數
        self.days_per_col = int(g_cfg.get("days_per_col", 7))
        self.days_per_col_auto = g_cfg.get("days_per_col_auto", True)
        # 建議引擎參數（A4/A3 橫向右側可用寬度，字元）
        rec = g_cfg.get("recommend", {})
        self.panel_widths = rec.get("panel_widths",
                                    {"A4_landscape": 120.0, "A3_landscape": 180.0})
        self.target_panel_width = rec.get("target_panel_width", 180.0)
        self.max_readable_dpc = rec.get("max_readable_days_per_col", 30)
        self.nice_steps = rec.get("nice_steps", [1, 2, 3, 5, 7, 10, 14, 21, 30])
        self.target_col_w = rec.get("target_col_width", 3.0)
        self.recommended_paper = None   # _recommend_granularity 填充
        self.recommendation = None      # 給 CLI/GUI 輸出用

    def _resolve_granularity(self, num_days: int) -> str:
        """Return 'day', 'week', or 'month' based on threshold + manual override."""
        if not self.gran_auto:
            return self.gran_manual
        if self.gran_manual and self.gran_manual != "auto":
            return self.gran_manual
        if num_days <= self.gran_thresholds.get("day_max", 120):
            return "day"
        if num_days <= self.gran_thresholds.get("week_max", 450):
            return "week"
        return "month"

    # ---- 顆粒度建議引擎（需求②：A4/A3 橫向自動建議）----
    def _compute_days_per_col(self, num_days: int) -> tuple:
        """Wrapper：列寬驅動壓縮計算。"""
        return compute_days_per_col(num_days,
                                    self.target_panel_width,
                                    self.nice_steps)

    def _recommend_granularity(self, num_days: int) -> dict:
        """比較 A4/A3 橫向，給出最優「日/列」建議。"""
        return recommend_granularity(num_days,
                                     self.panel_widths,
                                     self.nice_steps)

    def _build_slots(self):
        """Build time-slot array from self.dates.

        Sets:
          self.slots          — list of (label, start_date, end_date) tuples
          self.gantt_cols     — len(self.slots)
          self.granularity    — 'day' | 'week_grouped' | 'week' | 'compressed' | 'month'
          self.gran_label     — human-readable description for the subtitle line
          self.days_per_col   — compressed 模式生效值
        """
        self.granularity = self._resolve_granularity(self.num_days)

        if self.granularity == "day":
            self.slots = [(str(d.day), d, d) for d in self.dates]
            self.gantt_cols = len(self.slots)
            self.gran_label = "日"
            return

        if self.granularity == "week_grouped":
            # 每日一列，週分組欄由 _render_headers 處理；slot 同 day 模式
            self.slots = [(str(d.day), d, d) for d in self.dates]
            self.gantt_cols = len(self.slots)
            self.gran_label = "日（週分組）"
            return

        if self.granularity == "week":
            self.slots = []
            i = 0
            while i < len(self.dates):
                start = self.dates[i]
                days_to_sunday = 6 - start.weekday()
                end_idx = min(i + days_to_sunday, len(self.dates) - 1)
                end = self.dates[end_idx]
                iso_week = start.isocalendar()[1]
                year = start.isocalendar()[0]
                label = f"W{iso_week}"
                if i == 0 or year != self.dates[0].year:
                    label = f"{year}W{iso_week:02d}"
                self.slots.append((label, start, end))
                i = end_idx + 1
            self.gantt_cols = len(self.slots)
            self.gran_label = "週"
            return

        if self.granularity == "compressed":
            # 合併多天一列（壓縮顯示）：days_per_col 手動或自動
            if self.days_per_col_auto and self.num_days > 0:
                rec = self._recommend_granularity(self.num_days)
                self.recommendation = rec
                self.recommended_paper = rec["recommended"]["paper"]
                # 自動採用建議紙型對應嘅 days_per_col
                # 最小門檻：至少 2 日/列，確保用戶明確感知壓縮效果
                self.days_per_col = max(2, rec["recommended"]["days_per_col"])
                self.gantt_col_w = rec["recommended"].get("col_width", self.gran_col_w.get("compressed", 6.0))
            n = max(1, int(self.days_per_col))
            self.slots = []
            i = 0
            while i < len(self.dates):
                start = self.dates[i]
                end_idx = min(i + n - 1, len(self.dates) - 1)
                end = self.dates[end_idx]
                label = f"{start.month}/{start.day}"
                self.slots.append((label, start, end))
                i = end_idx + 1
            self.gantt_cols = len(self.slots)
            self.gran_label = f"壓縮（{n}日/列）"
            return

        # month granularity
        self.slots = []
        i = 0
        while i < len(self.dates):
            m = self.dates[i].month
            y = self.dates[i].year
            j = i
            while j < len(self.dates) and self.dates[j].month == m:
                j += 1
            label = f"{m}月" if y == self.dates[0].year else f"{y}年{m}月"
            self.slots.append((label, self.dates[i], self.dates[j - 1]))
            i = j
        self.gantt_cols = len(self.slots)
        self.gran_label = "月"

    def _load_i18n(self):
        i18n = self.sys_cfg.get("i18n", {})
        locale = i18n.get("current_locale", "zh_TW")
        self.trans = i18n.get("translations", {}).get(locale, {})

    def _load_output_cfg(self):
        out = self.sys_cfg.get("output", {})
        self.backup_enabled = out.get("backup_enabled", False)
        self.backup_count = out.get("backup_count", 5)
        self.version_suffix = out.get("version_suffix", "")
        self.export_pdf = out.get("export_pdf", False)

    def _load_validation(self):
        self.val_rules = self.sys_cfg.get("data_validation",
                                          self.sys_cfg.get("validation", {}))

    def _load_project(self):
        proj = self.task_cfg.get("project", {})
        self.project_name = proj.get("name", "工程項目")
        self.subtitle = proj.get("subtitle", "施工進度橫道圖")
        self.project_id = proj.get("project_id", "")
        self.period_text = proj.get("period_text", "")
        self.contractor = proj.get("contractor", "")
        self.start_date = self._parse_date(proj.get("start_date", "2026-07-27"))

        end_str = proj.get("end_date")
        if end_str:
            self.end_date = self._parse_date(end_str)
            self.num_days = (self.end_date - self.start_date).days + 1
        else:
            self.num_days = proj.get("num_days", 30)
            self.end_date = self.start_date + timedelta(days=self.num_days - 1)

        today_str = proj.get("today")
        if today_str:
            p = today_str.split("-")
            self.today = date(int(p[0]), int(p[1]), int(p[2]))
        else:
            self.today = date.today()

        self.legend = self.task_cfg.get("legend", [])
        self.notes = self.task_cfg.get("notes", [])

        # Determine info columns
        task_fields = self.task_cfg.get("task_fields", None)
        if task_fields:
            self.info_fields = self.INFO_BASE + [
                f for f in task_fields if f not in self.INFO_BASE
            ]
        else:
            self.info_fields = list(self.INFO_BASE)
        self.num_info = len(self.info_fields)

    # ---- Calendar ----

    def _build_calendar(self):
        incl_weekends = self.cal_type != "workdays"
        start_dt = self._to_dt(self.start_date)
        end_dt = self._to_dt(self.end_date)

        date_seq = TimeModule.generate_date_sequence(
            start_dt, end_dt,
            include_weekends=incl_weekends,
            holidays=[d.isoformat() for d in self.holidays],
        )
        self.dates = [d.date() if hasattr(d, "date") else d for d in date_seq]
        self.date_idx = {d: i for i, d in enumerate(self.dates)}
        self.gantt_cols = len(self.dates)   # raw day count, will be overridden by _build_slots
        self.total_workdays = sum(
            1 for d in self.dates
            if d.weekday() < 5 and d not in self.holidays
        )
        # Build granularity-aware time slots (may reduce gantt_cols)
        self._build_slots()

    # ---- Task processing ----

    def _process_tasks(self):
        self.sections = []
        self.all_tasks = []

        for sec in self.task_cfg.get("sections", []):
            sec_data = {"title": sec["title"], "tasks": []}
            for t in sec.get("tasks", []):
                task = {
                    "id": t["id"],
                    "name": t["name"],
                    "duration": t.get("duration", 1),
                    "start_day": t.get("start_day", 0),
                    "end_day": t.get("start_day", 0) + t.get("duration", 1) - 1,
                    "deps": t.get("deps", t.get("dependencies", [])),
                    "milestone": t.get("milestone", False),
                    "progress": t.get("progress"),
                    "category": t.get("category", "A"),
                    "color": t.get("color",
                                   self.stage_colors.get(
                                       t.get("category", "A"), "2E86AB")),
                    "material": t.get("material", ""),
                    "remark": t.get("remark", ""),
                    "critical": False,
                }
                task["start_date"] = self._day_to_date(task["start_day"])
                task["end_date"] = self._day_to_date(task["end_day"])
                sec_data["tasks"].append(task)
                self.all_tasks.append(task)
            self.sections.append(sec_data)

    def _validate(self):
        self.warnings = validate_tasks(self.all_tasks, self.val_rules)

    def _compute_cp(self):
        cp_cfg = self.task_cfg.get("critical_path", [])
        dep_cfg = self.sys_cfg.get("dependency", {})
        auto = dep_cfg.get("auto_calculate_critical_path", True)

        if cp_cfg:
            self.cp_ids = set(cp_cfg)
        elif auto:
            self.cp_ids = self._auto_cp()
        else:
            self.cp_ids = set()

        for t in self.all_tasks:
            t["critical"] = t["id"] in self.cp_ids

    def _auto_cp(self) -> set:
        tasks = {t["id"]: t for t in self.all_tasks}
        memo = {}

        def end_of(tid):
            if tid in memo:
                return memo[tid]
            t = tasks[tid]
            if not t["deps"]:
                val = t["end_day"]
            else:
                val = max(
                    (end_of(d) for d in t["deps"] if d in tasks),
                    default=t["start_day"] - 1,
                ) + t["duration"]
            memo[tid] = val
            return val

        for t in self.all_tasks:
            end_of(t["id"])

        if not memo:
            return set()
        max_end = max(memo.values())
        return {tid for tid, e in memo.items() if e == max_end}

    # ---- Rendering pipeline ----

    def _render(self):
        tc = self.colors.get("GRID_BOLD", "1E3B4A")
        self.ws = self.wb.active
        self.ws.title = "施工進度橫道圖"
        self.ws.sheet_properties.tabColor = tc

        self._render_title()
        self._render_headers()
        self._render_body()
        self._render_today_marker()
        self._render_legend()
        self._render_notes()
        self._set_widths()
        self._set_freeze()
        self._set_print()

        if self.backup_enabled:
            self._do_backup()

        os.makedirs(os.path.dirname(os.path.abspath(self.output)), exist_ok=True)
        self.wb.save(self.output)

    # ---- Style helpers ----

    def _font(self, size=10, bold=False, color="000000", name=None):
        return Font(name=name or self.font_name, size=size, bold=bold, color=color)

    def _fill(self, color):
        return PatternFill("solid", fgColor=color)

    def _border(self, color=None, style="thin"):
        c = color or self.colors.get("GRID", "C5D1DF")
        s = Side(style, color=c)
        return Border(left=s, right=s, top=s, bottom=s)

    def _border_thick_bottom(self):
        gc = self.colors.get("GRID", "C5D1DF")
        bc = self.colors.get("GRID_BOLD", "1E3B4A")
        return Border(
            left=Side("thin", color=gc), right=Side("thin", color=gc),
            top=Side("thin", color=gc), bottom=Side("medium", color=bc),
        )

    @staticmethod
    def _align(h="center"):
        return Alignment(horizontal=h, vertical="center", wrap_text=True)

    def _set_cell(self, row, col, value=None,
                  font=None, fill=None, align=None, border=None):
        c = self.ws.cell(row=row, column=col)
        if value is not None:
            c.value = value
        if font:
            c.font = font
        if fill:
            c.fill = fill
        if align:
            c.alignment = align
        if border:
            c.border = border
        return c

    def _merge(self, r1, c1, r2, c2, value=None,
               font=None, fill=None, align=None, border=None):
        self.ws.merge_cells(start_row=r1, start_column=c1,
                            end_row=r2, end_column=c2)
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                cell = self.ws.cell(row=r, column=c)
                if font:
                    cell.font = font
                if fill:
                    cell.fill = fill
                if align:
                    cell.alignment = align
                if border:
                    cell.border = border
        if value is not None:
            self.ws.cell(row=r1, column=c1).value = value

    def _info_label(self, field_key: str) -> str:
        return tr(self.trans, field_key,
                  self.INFO_LABELS_FALLBACK.get(field_key, field_key))

    def _fmt_date(self, d: date) -> str:
        """Format date for info-column display.

        Within the same year as project_start → "M/D" (compact).
        Different year → "YYYY-M-D" to avoid ambiguity.
        """
        if d.year == self.start_date.year:
            return f"{d.month}/{d.day}"
        return f"{d.year}/{d.month}/{d.day}"

    # ---- Title ----

    def _render_title(self):
        R_TITLE, R_SUB, R_SPC = 1, 2, 3
        gc = self.num_info + self.gantt_cols
        hdr_bg = self.colors.get("HEADER_BG", "1E3B4A")
        hdr_fg = self.colors.get("HEADER_FG", "FFFFFF")

        # Row 1
        self._merge(R_TITLE, 1, R_TITLE, gc,
                    f"{self.project_name}  {self.subtitle}",
                    self._font(self.font_sizes.get("title", 14), True, hdr_fg),
                    self._fill(hdr_bg), self._align(),
                    self._border_thick_bottom())
        self.ws.row_dimensions[R_TITLE].height = self.row_heights.get(
            "title_row", 38)

        # Row 2
        d0, d1 = self.dates[0], self.dates[-1]
        period_text = getattr(self, "period_text", "")
        id_part = f"{tr(self.trans, 'project_name', '工程編號')}: {self.project_id}" if self.project_id else ""
        contractor_part = f"  |  {tr(self.trans, 'constructor', '承建商')}: {self.contractor}" if self.contractor else ""
        if period_text:
            period_part = f"{tr(self.trans, 'period', '施工期間')}：{period_text}"
        else:
            period_part = (
                f"{tr(self.trans, 'period', '工期')}：{self.num_days}{tr(self.trans, 'calendar_day', '日曆天')}"
                f"（{d0.month}月{d0.day}日～{d1.month}月{d1.day}日）"
            )
        parts = [id_part, period_part,
                 f"假設工程委托於：{d0.month}月{d0.day}日（{WEEKDAY_CN[d0.weekday()]}）開工",
                 f"時間刻度：{self.gran_label}"]
        if contractor_part.strip():
            parts.insert(1, contractor_part.strip())
        sub = "  |  ".join(p for p in parts if p)
        font_color = self.colors.get("FONT_COLOR", "2C3E50")
        self._merge(R_SUB, 1, R_SUB, gc, sub,
                    self._font(9, False, font_color),
                    self._fill(self.colors.get("INFO_BG", "F0F4FA")),
                    self._align(), self._border())
        self.ws.row_dimensions[R_SUB].height = 24

        # Row 3 spacer
        self.ws.row_dimensions[R_SPC].height = 6
        for c in range(1, gc + 1):
            self.ws.cell(row=R_SPC, column=c).border = Border()

    # ---- Headers ----

    def _render_headers(self):
        g0 = self.num_info + 1
        C = self.colors
        hdr_bg = C.get("HEADER_BG", "1E3B4A")
        info_bg = C.get("INFO_BG", "F0F4FA")
        we_bg = C.get("WEEKEND_BG", "E8E8E8")
        today_bg = C.get("TODAY_COL", "FCEFE3")
        today_ln = C.get("TODAY_LINE", "D9534F")
        font_c = C.get("FONT_COLOR", "2C3E50")
        thin = self._border()
        thick = self._border_thick_bottom()

        # ---- Info area (rows 4-6, cols A..num_info) — no merge ----
        for r in (4, 5, 6):
            for c in range(1, self.num_info + 1):
                cell = self.ws.cell(row=r, column=c)
                cell.fill = self._fill(info_bg)
                cell.border = thin

        # Row 5: column labels
        for ci, field in enumerate(self.info_fields, 1):
            cell = self.ws.cell(row=5, column=ci)
            cell.value = self._info_label(field)
            cell.font = self._font(9, True, font_c)
            cell.fill = self._fill(info_bg)
            cell.alignment = self._align()
            cell.border = thick

        # ---- Top header row 4: grouping bands (gantt area) ----
        month_alt = lighten(hdr_bg, 0.5)
        if self.granularity == "week_grouped":
            # 每周一栏：合併同一 ISO 週嘅日列，標籤「第N週 起–止」
            wk_colors = [hdr_bg, month_alt]
            wi = 0
            i = 0
            while i < len(self.slots):
                s_start = self.slots[i][1]
                iy, iw = s_start.isocalendar()[0], s_start.isocalendar()[1]
                j = i
                while j < len(self.slots):
                    js = self.slots[j][1]
                    if js.isocalendar()[0] != iy or js.isocalendar()[1] != iw:
                        break
                    j += 1
                s_end = self.slots[j - 1][2]
                wi += 1
                wk_label = (f"第{wi}週 {s_start.month}/{s_start.day}"
                            f"–{s_end.month}/{s_end.day}")
                wc = wk_colors[(wi - 1) % 2]
                self._merge(4, g0 + i, 4, g0 + j - 1, wk_label,
                            self._font(10, True, C.get("HEADER_FG", "FFFFFF")),
                            self._fill(wc), self._align(), thin)
                i = j
        else:
            # 月份分組（day / week / compressed / month 共用）
            m_colors = [hdr_bg, month_alt]
            mi = 0
            i = 0
            while i < len(self.slots):
                label, s_start, s_end = self.slots[i]
                # Group consecutive slots by month for the top band
                m = s_start.month
                y = s_start.year
                j = i
                while j < len(self.slots) and self.slots[j][1].month == m:
                    j += 1
                mc = m_colors[mi % 2]
                m_label = f"{m}月" if y == self.dates[0].year else f"{y}年{m}月"
                self._merge(4, g0 + i, 4, g0 + j - 1, m_label,
                            self._font(10, True, C.get("HEADER_FG", "FFFFFF")),
                            self._fill(mc), self._align(), thin)
                mi += 1
                i = j

        # ---- Row 5: slot labels (day number / week label / month label) ----
        for idx, (label, s_start, s_end) in enumerate(self.slots):
            col = g0 + idx
            is_today = s_start <= self.today <= s_end
            if is_today:
                bg, fg = today_bg, today_ln
            else:
                bg, fg = hdr_bg, C.get("HEADER_FG", "FFFFFF")
            cell = self.ws.cell(row=5, column=col)
            cell.value = label
            cell.font = self._font(9, True, fg)
            cell.fill = self._fill(bg)
            cell.alignment = self._align()
            cell.border = thin

        # ---- Row 6: sub-labels ----
        if self.granularity in ("day", "week_grouped"):
            for idx, (label, s_start, s_end) in enumerate(self.slots):
                col = g0 + idx
                d = s_start
                is_we = d.weekday() >= 5 or d in self.holidays
                is_today = d == self.today
                if is_today:
                    fg, bg = today_ln, today_bg
                elif is_we:
                    fg, bg = C.get("DEPENDENCY_MARKER", "C0503D"), we_bg
                else:
                    fg, bg = font_c, "FFFFFF"
                cell = self.ws.cell(row=6, column=col)
                cell.value = WEEKDAY_CN[d.weekday()]
                cell.font = self._font(8, False, fg)
                cell.fill = self._fill(bg)
                cell.alignment = self._align()
                cell.border = thin
        elif self.granularity == "compressed":
            # 顯示每列範圍止日，配合 row5 起日成「起–止」
            for idx, (label, s_start, s_end) in enumerate(self.slots):
                col = g0 + idx
                cell = self.ws.cell(row=6, column=col)
                cell.value = f"–{s_end.month}/{s_end.day}"
                cell.font = self._font(8, False, font_c)
                cell.fill = self._fill("FFFFFF")
                cell.alignment = self._align()
                cell.border = thin
        else:
            for idx in range(len(self.slots)):
                col = g0 + idx
                cell = self.ws.cell(row=6, column=col)
                cell.font = self._font(8, False, font_c)
                cell.fill = self._fill("FFFFFF")
                cell.alignment = self._align()
                cell.border = thin

        # Row heights
        self.ws.row_dimensions[4].height = self.row_heights.get("scale_month", 22)
        self.ws.row_dimensions[5].height = self.row_heights.get("scale_day", 22)
        self.ws.row_dimensions[6].height = self.row_heights.get("scale_weekday", 18)

    # ---- Body ----

    def _render_body(self):
        row = 7
        gc = self.num_info + self.gantt_cols
        sec_fill = self._fill(self.colors.get("ALTERNATE_ODD", "F4F7FB"))

        for si, sec in enumerate(self.sections):
            # Section header
            self._merge(row, 1, row, gc, sec["title"],
                        self._font(10, True,
                                   self.colors.get("FONT_COLOR", "2C3E50")),
                        sec_fill, self._align("left"),
                        self._border_thick_bottom())
            self.ws.row_dimensions[row].height = self.row_heights.get(
                "header_row", 26)
            row += 1

            is_odd = si % 2 == 1
            for task in sec["tasks"]:
                self._render_task(task, row, is_odd)
                row += 1

    def _render_task(self, task, row, is_odd_sec):
        g0 = self.num_info + 1
        C = self.colors
        is_cp = task["critical"]
        color = task["color"]

        self.ws.row_dimensions[row].height = self.row_heights.get("task_row", 26)

        # Row background
        if is_cp:
            row_fill = self._fill(C.get("TODAY_COL", "FCEFE3"))
        elif is_odd_sec:
            row_fill = self._fill(C.get("ALTERNATE_ODD", "F4F7FB"))
        else:
            row_fill = self._fill(C.get("ALTERNATE_EVEN", "FFFFFF"))

        thin = self._border()
        fc = C.get("FONT_COLOR", "2C3E50")

        # Info columns
        for ci, field in enumerate(self.info_fields, 1):
            val = None
            fn = self._font(9, False, fc)
            al = self._align()

            if field == "seq":
                val = task["id"]
                fn = self._font(9, True,
                                C.get("DEPENDENCY_MARKER", "C0503D")
                                if is_cp else "546E7A")
            elif field == "name":
                val = task["name"]
                fn = self._font(10, False, fc)
                al = self._align("left")
            elif field == "duration":
                val = task["duration"]
            elif field == "start":
                val = self._fmt_date(task["start_date"])
                fn = self._font(8, False, "546E7A")
            elif field == "end":
                val = self._fmt_date(task["end_date"])
                fn = self._font(8, False, "546E7A")
            elif field == "predecessor":
                dep_ids = task.get("deps", [])
                val = ", ".join(str(d) for d in dep_ids) if dep_ids else ""
                fn = self._font(8, False, "78909C")
            elif field == "material":
                val = task.get("material", "")
                fn = self._font(8, False, "78909C")
            elif field == "remark":
                val = task.get("remark", "")
                fn = self._font(8, False, "78909C")
            elif field == "category":
                val = task.get("category", "")
                fn = self._font(8, True, color)

            self._set_cell(row, ci, val,
                           font=fn, fill=row_fill, align=al, border=thin)

        # ---- Gantt bars (slot-based) ----
        # Compute task's slot range from day indices
        task_start_date = task["start_date"]
        task_end_date = task["end_date"]
        start_slot = None
        end_slot = None
        for si, (label, s_start, s_end) in enumerate(self.slots):
            if s_start <= task_start_date <= s_end:
                start_slot = si
            if s_start <= task_end_date <= s_end:
                end_slot = si
        if start_slot is None:
            start_slot = 0
        if end_slot is None:
            end_slot = len(self.slots) - 1

        span = end_slot - start_slot

        for si in range(len(self.slots)):
            col = g0 + si
            s_start, s_end = self.slots[si][1], self.slots[si][2]

            if si < start_slot or si > end_slot:
                # Outside bar: white or weekend background
                is_we = s_start.weekday() >= 5 and self.granularity in ("day", "week_grouped")
                is_hol = s_start in self.holidays and self.granularity in ("day", "week_grouped")
                if is_hol:
                    bg = self._fill(C.get("DARK_WEEKEND_BG", "D0D0D0"))
                elif is_we:
                    bg = self._fill(C.get("WEEKEND_BG", "E8E8E8"))
                else:
                    bg = row_fill
                self._set_cell(row, col, None, fill=bg, border=thin)
                continue

            # Inside bar
            if task["milestone"] and si == start_slot:
                self._set_cell(row, col, "\u25C6",
                               self._font(12, True, color),
                               row_fill, self._align(), thin)
            else:
                pos = si - start_slot
                if span == 0:
                    bf = self._fill(color)
                elif pos == 0:
                    bf = self._fill(lighten(color, 0.35))
                elif pos == span:
                    bf = self._fill(darken(color, 0.18))
                else:
                    bf = self._fill(color)

                self._set_cell(row, col, None, fill=bf, border=thin)


                # Progress indicator
                prog = task.get("progress")
                if prog and prog > 0:
                    prog_slots = int((span + 1) * prog / 100)
                    if pos < prog_slots:
                        self.ws.cell(row=row, column=col).fill = self._fill(
                            darken(color, 0.12))

    # ---- Today marker ----

    def _render_today_marker(self):
        """Highlight today's slot column with a red accent."""
        # Find today's slot index
        today_slot = None
        for si, (label, s_start, s_end) in enumerate(self.slots):
            if s_start <= self.today <= s_end:
                today_slot = si
                break
        if today_slot is None:
            return

        g0 = self.num_info + 1
        tc = g0 + today_slot
        tl = C_today = self.colors.get("TODAY_LINE", "D9534F")

        red_left = Border(
            left=Side("medium", color=tl),
            right=Side("thin", color=self.colors.get("GRID", "C5D1DF")),
            top=Side("thin", color=self.colors.get("GRID", "C5D1DF")),
            bottom=Side("thin", color=self.colors.get("GRID", "C5D1DF")),
        )

        # Highlight today's header cells
        self.ws.cell(row=5, column=tc).font = self._font(9, True, "FFFFFF")
        self.ws.cell(row=5, column=tc).fill = self._fill(tl)
        self.ws.cell(row=6, column=tc).font = self._font(8, True, tl)

        # Red left border on body cells
        body_end = 6
        for sec in self.sections:
            body_end += 1 + len(sec["tasks"])
        for r in range(7, body_end + 1):
            self.ws.cell(row=r, column=tc).border = red_left

    # ---- Legend ----

    def _render_legend(self):
        if not self.legend:
            return

        gc = self.num_info + self.gantt_cols
        row = self.ws.max_row + 2
        fc = self.colors.get("FONT_COLOR", "2C3E50")

        self._merge(row, 1, row, gc,
                    f"{tr(self.trans, 'legend_label', '圖例')} Legend",
                    self._font(10, True, fc),
                    self._fill(self.colors.get("INFO_BG", "F0F4FA")),
                    self._align("left"), self._border())
        self.ws.row_dimensions[row].height = 24
        row += 1

        thin = self._border()
        per_row = 3
        for i in range(0, len(self.legend), per_row):
            items = self.legend[i:i + per_row]
            for j, item in enumerate(items):
                sc = item["color"]
                sl = item["label"]
                c1 = 1 + j * 2
                c2 = 2 + j * 2
                self._set_cell(row, c1, None,
                               fill=self._fill(sc), border=thin)
                self._set_cell(row, c2, sl,
                               font=self._font(9, False, fc),
                               align=self._align("left"),
                               border=thin)
            self.ws.row_dimensions[row].height = 20
            row += 1

    # ---- Notes ----

    def _render_notes(self):
        if not self.notes:
            return

        gc = self.num_info + self.gantt_cols
        row = self.ws.max_row + 2

        self._merge(row, 1, row, gc,
                    f"{tr(self.trans, 'critical_path_label', '關鍵路徑')}分析",
                    self._font(10, True, "E65100"),
                    self._fill("FFF8E1"), self._align("left"), self._border())
        self.ws.row_dimensions[row].height = 24
        row += 1

        nf = self._font(9, False, "5D4037")
        nfill = self._fill("FFF8E1")
        nal = self._align("left")
        nbdr = self._border()

        for note in self.notes:
            self._merge(row, 1, row, gc, note,
                        nf, nfill, nal, nbdr)
            self.ws.row_dimensions[row].height = 22
            row += 1

    # ---- Widths, freeze, print ----

    def _set_widths(self):
        cw = self.info_col_widths
        col_letters = ["A", "B", "C", "D", "E", "F", "G", "H"]
        defaults = {
            "seq": 5, "name": 36, "duration": 6, "start": 8, "end": 8,
            "predecessor": 10, "material": 14, "remark": 16, "category": 8,
        }
        for i, field in enumerate(self.info_fields):
            letter = col_letters[i] if i < len(col_letters) else get_column_letter(i + 1)
            w = cw.get(field, defaults.get(field, 10))
            self.ws.column_dimensions[letter].width = w

        # Gantt columns: auto-select width by granularity
        # compressed 模式優先用推薦引擎算出的 col_width
        if self.granularity == "compressed" and hasattr(self, 'gantt_col_w'):
            gantt_w = self.gantt_col_w
        else:
            gantt_w = self.gran_col_w.get(self.granularity,
                      self.sys_cfg.get("gantt_layout", {}).get("gantt_column_width", 4.5))
        g0 = self.num_info + 1
        for i in range(self.gantt_cols):
            self.ws.column_dimensions[get_column_letter(g0 + i)].width = gantt_w

    def _set_freeze(self):
        self.ws.freeze_panes = self.freeze_cfg

    def _set_print(self):
        sp = self.ws.sheet_properties
        sp.pageSetUpPr.fitToPage = True
        self.ws.page_setup.orientation = "landscape"

        # 紙型優先級（2026-07-23 修復）：
        # 1) 尊重 config_v2.json 的 print.paper_size（8=A3, 9=A4）——
        #    不再無視明確設定而強制 A4；這是 GanttChart Pro 標準（用戶要求 A3 橫向）。
        # 2) 有推薦紙型且 config 未明設時，按推薦。
        # 3) 否則自動：≤60 列→A4，否則 A3。
        cfg_paper = self.sys_cfg.get("print", {}).get("paper_size", None)
        if cfg_paper in (8, 9):
            paper = cfg_paper
        elif self.recommended_paper:
            paper = 9 if self.recommended_paper == "A4" else 8
        elif self.gantt_cols <= 60:
            paper = 9   # A4
        else:
            paper = 8   # A3
        self.ws.page_setup.paperSize = paper
        self.ws.page_setup.fitToWidth = self.sys_cfg.get("print", {}).get("fit_to_pages_wide", 1)
        self.ws.page_setup.fitToHeight = 0  # unlimited pages tall

        # 重複列（每頁重印標題/信息/表頭）：GanttChart Pro 標準 1:6
        # （2026-07-23 修復：原引擎從未設定 print titles）
        try:
            self.ws.print_title_rows = "1:6"
        except Exception:
            pass

        # Margins (inches) - read from config if available
        mg = self.sys_cfg.get("print", {}).get("margins", {})
        self.ws.page_margins.left = mg.get("left", 0.4)
        self.ws.page_margins.right = mg.get("right", 0.4)
        self.ws.page_margins.top = mg.get("top", 0.5)
        self.ws.page_margins.bottom = mg.get("bottom", 0.5)
        self.ws.page_margins.header = mg.get("header", 0.2)
        self.ws.page_margins.footer = mg.get("footer", 0.2)

        pr = self.sys_cfg.get("print", {})
        if "header_left" in pr:
            self.ws.oddHeader.left.text = pr["header_left"]
        if "footer_right" in pr:
            self.ws.oddFooter.right.text = pr["footer_right"]

        # 打印選項（2026-08-01 修復）：config 內 print_gridlines /
        # print_headings 一直係死設定，引擎從未套用；交付文件唔應該
        # 印網格線同行號列標。
        self.ws.print_options.gridLines = bool(pr.get("print_gridlines", False))
        self.ws.print_options.headings = bool(pr.get("print_headings", False))
        self.ws.print_options.horizontalCentered = bool(
            pr.get("horizontally_centered", True))

        # 打印範圍：明確界定，避免 Excel 誤判尾部空白列而多印一頁
        try:
            self.ws.print_area = (
                f"A1:{get_column_letter(self.ws.max_column)}{self.ws.max_row}")
        except Exception:
            pass

    # ---- Helpers ----

    @staticmethod
    def _parse_date(s: str) -> date:
        p = s.split("-")
        return date(int(p[0]), int(p[1]), int(p[2]))

    @staticmethod
    def _to_dt(d: date):
        from datetime import datetime
        return datetime(d.year, d.month, d.day)

    def _day_to_date(self, day_offset: int) -> date:
        if 0 <= day_offset < len(self.dates):
            return self.dates[day_offset]
        return self.start_date + timedelta(days=day_offset)

    # ---- Backup ----

    def _do_backup(self):
        if not os.path.exists(self.output):
            return
        base, ext = os.path.splitext(self.output)
        for i in range(self.backup_count - 1, 0, -1):
            src = f"{base}_v{i}{ext}"
            dst = f"{base}_v{i + 1}{ext}"
            if os.path.exists(src):
                shutil.move(src, dst)
        shutil.copy2(self.output, f"{base}_v1{ext}")


# ================================================================
#  Entry Point
# ================================================================

def compute_days_per_col(num_days, right_panel_width=180.0, nice_steps=None):
    """列寬驅動算法：在可讀列寬區間內找最佳「日/列」。

    與 V15 對齊：
      - 計算 col_width = right_panel_width / cols
      - 優先找 col_width ∈ [2.0, 15.0] 的間隔（可讀範圍）
      - 若無，找最接近 6.0pt 的間隔

    Args:
        num_days: 項目總天數
        right_panel_width: 右側甘特區域總寬度（字元），預設 180（A3橫向約180字元）
        nice_steps: 候選間隔列表，預設 [1,2,3,5,7,10,14,21,30]

    Returns:
        (interval_days, num_cols, col_width)  日/列, 總列數, 每列寬度
    """
    if nice_steps is None:
        nice_steps = [1, 2, 3, 5, 7, 10, 14, 21, 30]

    # 階段 1：優先找列寬在 2.0~15.0 可讀範圍內的間隔
    for interval in nice_steps:
        cols = (num_days + interval - 1) // interval
        cw = right_panel_width / cols
        if 2.0 <= cw <= 15.0:
            return interval, cols, cw

    # 階段 2：若全都不在範圍內，找最接近 6.0pt 的
    best, best_diff = nice_steps[-1], float('inf')
    for interval in nice_steps:
        cols = (num_days + interval - 1) // interval
        cw = right_panel_width / cols
        if abs(cw - 6.0) < best_diff:
            best_diff = abs(cw - 6.0)
            best = interval
    cols = (num_days + best - 1) // best
    return best, cols, right_panel_width / cols


def recommend_granularity(num_days, right_panel_widths=None, nice_steps=None):
    """比較 A4 橫向 / A3 橫向，給出最優「日/列」建議。

    採用列寬驅動算法，而非紙型容量減法。
    A4 橫向可用約 120 字元，A3 橫向約 180 字元。
    """
    if right_panel_widths is None:
        right_panel_widths = {"A4_landscape": 120.0, "A3_landscape": 180.0}
    if nice_steps is None:
        nice_steps = [1, 2, 3, 5, 7, 10, 14, 21, 30]

    w_a4 = right_panel_widths.get("A4_landscape", 120.0)
    w_a3 = right_panel_widths.get("A3_landscape", 180.0)

    dpc_a4, cols_a4, cw_a4 = compute_days_per_col(num_days, w_a4, nice_steps)
    dpc_a3, cols_a3, cw_a3 = compute_days_per_col(num_days, w_a3, nice_steps)

    # 優先 A4，若 A4 列數太多（>120列擠不下）改用 A3
    if cols_a4 <= 120:
        rec = {"paper": "A4", "days_per_col": dpc_a4, "cols": cols_a4, "col_width": cw_a4}
    else:
        rec = {"paper": "A3", "days_per_col": dpc_a3, "cols": cols_a3, "col_width": cw_a3}

    return {
        "num_days": num_days,
        "A4": {"paper": "A4", "days_per_col": dpc_a4, "cols": cols_a4, "col_width": cw_a4},
        "A3": {"paper": "A3", "days_per_col": dpc_a3, "cols": cols_a3, "col_width": cw_a3},
        "recommended": rec,
    }


def _parse_date_simple(s):
    s = str(s).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"無法解析日期: {s}")


def _project_num_days(proj):
    start = proj.get("start_date")
    end = proj.get("end_date")
    if start and end:
        try:
            sd = _parse_date_simple(start)
            ed = _parse_date_simple(end)
            return (ed - sd).days + 1
        except ValueError:
            pass
    return int(proj.get("num_days", 30))


def _print_recommendation(rec):
    a4 = rec["A4"]; a3 = rec["A3"]; r = rec["recommended"]
    print("=== 時間顆粒度建議（壓縮模式 / A4·A3 橫向）===")
    print(f"項目總日數 : {rec['num_days']} 日")
    print(f"A4 橫向    : {a4['days_per_col']} 日/列 → 共 {a4['cols']} 列")
    print(f"A3 橫向    : {a3['days_per_col']} 日/列 → 共 {a3['cols']} 列")
    print(f"★ 建議     : {r['paper']} 橫向，{r['days_per_col']} 日/列（共 {r['cols']} 列）")
    print("  生成: python gen_gantt.py 輸入.xlsx 輸出.xlsx --granularity compressed")
    print(f"  或手動: --days-per-col {r['days_per_col']}")


def main():
    import argparse
    parser = argparse.ArgumentParser(
        description="施工進度橫道圖生成器（day / week_grouped / week / compressed / month）")
    parser.add_argument("task_path", nargs="?",
                        help="EDF.xlsx 或 gantt_config.json；留空啟動 GUI")
    parser.add_argument("out_path", nargs="?", default=None,
                        help="輸出 .xlsx（可選）")
    parser.add_argument("--project", default="", help="覆寫工程名稱")
    parser.add_argument("--subtitle", default="", help="覆寫副標題")
    parser.add_argument("--granularity", default=None,
                        choices=["auto", "day", "week_grouped", "week",
                                 "compressed", "month"],
                        help="時間顆粒度；compressed 合併多天一列並自動建議")
    parser.add_argument("--days-per-col", type=int, default=None,
                        help="compressed 手動指定「日/列」（唔設則自動建議）")
    parser.add_argument("--scheme", default=None, help="配色方案")
    parser.add_argument("--suggest", action="store_true",
                        help="只列印 A4/A3 橫向「日/列」建議，唔生成圖")
    args = parser.parse_args()

    script_dir = os.path.dirname(os.path.abspath(__file__))

    # ---- No task_path = launch GUI ----
    if not args.task_path:
        try:
            from gui_gantt import launch_gui
            launch_gui()
            return
        except ImportError:
            print("GUI 模組未找到，請提供輸入檔案")
            return

    task_path = args.task_path
    task_ext = os.path.splitext(task_path)[1].lower()
    if task_ext in (".xlsx", ".xlsm", ".xls"):
        from edf_importer import import_edf
        task_config = import_edf(task_path,
                                 project_name=args.project,
                                 subtitle=args.subtitle)
    else:
        with open(task_path, "r", encoding="utf-8") as f:
            task_config = json.load(f)

    # ---- 載入系統配置 ----
    sys_config = {}
    sys_path = os.path.join(script_dir, "config_v2.json")
    if os.path.exists(sys_path):
        with open(sys_path, "r", encoding="utf-8") as f:
            sys_config = json.load(f)

    # ---- 套用 CLI 顆粒度選項 ----
    g_cfg = sys_config.setdefault("granularity", {})
    if args.granularity:
        g_cfg["manual"] = args.granularity
        g_cfg["auto"] = (args.granularity == "auto")
    if args.days_per_col:
        g_cfg["days_per_col"] = args.days_per_col
        g_cfg["days_per_col_auto"] = False

    # ---- 只建議模式 ----
    if args.suggest:
        rec_cfg = g_cfg.get("recommend", {})
        rec = recommend_granularity(
            _project_num_days(task_config.get("project", {})),
            right_panel_widths=rec_cfg.get("panel_widths",
                        {"A4_landscape": 120.0, "A3_landscape": 180.0}),
            nice_steps=rec_cfg.get("nice_steps", [1, 2, 3, 5, 7, 10, 14, 21, 30])
        )
        _print_recommendation(rec)
        return

    styles_config = {}
    sty_path = os.path.join(script_dir, "gantt_styles.json")
    if os.path.exists(sty_path):
        with open(sty_path, "r", encoding="utf-8") as f:
            styles_config = json.load(f)

    if args.out_path:
        out_path = args.out_path
    else:
        pname = task_config.get("project", {}).get("name", "工程項目")
        out_path = os.path.join(script_dir, f"{pname}-施工進度橫道圖.xlsx")

    if args.scheme:
        sys_config["active_scheme"] = args.scheme

    print(f"Config  : {os.path.basename(task_path)}")
    print(f"Gran    : {g_cfg.get('manual', 'auto')}")
    print(f"Output  : {out_path}")

    chart = GanttChart(task_config, out_path, sys_config, styles_config)

    if chart.recommendation:
        _print_recommendation(chart.recommendation)
    if chart.warnings:
        for w in chart.warnings:
            print(f"  [!] {w}")
    print(f"Done    -> {out_path}")


if __name__ == "__main__":
    main()
